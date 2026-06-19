from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "output" / "lead6_report.xlsx"


SUMMARY = {
    "deposits_total": 5632.00,
    "expenses_total": 4463.26,
    "transfer_fees_total": 1.73,
    "net_total": 1168.74,
    "deposits_count": 7,
    "expenses_count": 4,
}


ROWS = [
    {
        "date": datetime(2026, 6, 2),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 1150.00,
        "reference": "عملية تحويل داخلية",
        "note": "إيداع بنكي داخلي",
    },
    {
        "date": datetime(2026, 6, 3),
        "move_type": "مصروف",
        "expense_type": "فاتورة 12",
        "amount": 3477.39,
        "note": "حوالة فورية صادرة لتقنية/سداد فاتورة",
    },
    {
        "date": datetime(2026, 6, 3),
        "move_type": "مصروف",
        "expense_type": "رسوم فاتورة 12",
        "amount": 1.15,
        "note": "رسوم حوالة فورية صادرة لتقنية/سداد فاتورة",
    },
    {
        "date": datetime(2026, 6, 4),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 147.00,
        "reference": "20260604SAINMAINMA2BXXX11042505197",
        "note": "حوالات فورية واردة",
    },
    {
        "date": datetime(2026, 6, 5),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 575.00,
        "reference": "20260605SANCBKNCBK6B82410258558827",
        "note": "حوالات فورية واردة",
    },
    {
        "date": datetime(2026, 6, 5),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 450.00,
        "reference": "20260607SAINMAINMA2BXXX11434519126",
        "note": "حوالات فورية واردة",
    },
    {
        "date": datetime(2026, 6, 7),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 1150.00,
        "reference": "عملية تحويل داخلية",
        "note": "إيداع بنكي داخلي",
    },
    {
        "date": datetime(2026, 6, 8),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 900.00,
        "reference": "20260610SAINMAINMA2BXXX12244500377",
        "note": "حوالات فورية واردة",
    },
    {
        "date": datetime(2026, 6, 9),
        "move_type": "إيداع",
        "expense_type": "",
        "amount": 1260.00,
        "reference": "20260611SAINMAINMA2BXXX12155103618",
        "note": "حوالات فورية واردة",
    },
    {
        "date": datetime(2026, 6, 10),
        "move_type": "مصروف",
        "expense_type": "فاتورة 13",
        "amount": 984.14,
        "note": "حوالة فورية صادرة لتقنية/فاتورة",
    },
    {
        "date": datetime(2026, 6, 10),
        "move_type": "مصروف",
        "expense_type": "رسوم فاتورة 13",
        "amount": 0.58,
        "note": "رسوم حوالة فورية صادرة لتقنية/فاتورة",
    },
]


def style_range(ws, start_row, end_row, start_col, end_col, fill=None, font=None, border=None, alignment=None):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def build_statement_sheet():
    wb = load_workbook(SOURCE)
    if "كشف الحساب" in wb.sheetnames:
        del wb["كشف الحساب"]
    ws = wb.create_sheet("كشف الحساب")
    ws.sheet_view.showGridLines = False

    dark = PatternFill("solid", fgColor="5B4A47")
    soft = PatternFill("solid", fgColor="F7F1EA")
    green = PatternFill("solid", fgColor="EEF7F0")
    red = PatternFill("solid", fgColor="FFF1F0")
    header_fill = PatternFill("solid", fgColor="2F2628")
    white_font = Font(color="FFFFFF", bold=True, size=15)
    bold_font = Font(bold=True, color="2F2628")
    label_font = Font(bold=True, color="5B4A47")
    body_font = Font(color="2A2425")
    thin = Side(style="thin", color="E1D4CA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.merge_cells("A1:F1")
    ws["A1"] = "كشف الحساب الجاري"
    ws["A1"].fill = dark
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = right

    ws.merge_cells("A2:F2")
    ws["A2"] = "المصدر: كشف حساب جاري PDF"
    ws["A2"].font = Font(color="7C6B68", italic=True, size=11)
    ws["A2"].alignment = right

    summary_items = [
        ("إجمالي الإيداعات", SUMMARY["deposits_total"], "B4"),
        ("إجمالي المصاريف", SUMMARY["expenses_total"], "B5"),
        ("صافي الحركة", SUMMARY["net_total"], "B6"),
        ("عدد الإيداعات", SUMMARY["deposits_count"], "E4"),
        ("عدد المصاريف", SUMMARY["expenses_count"], "E5"),
        ("الفترة", "2026/06/01 - 2026/06/10", "E6"),
    ]

    ws.merge_cells("A3:B3")
    ws["A3"] = "ملخص الكشف"
    ws["A3"].fill = header_fill
    ws["A3"].font = white_font
    ws["A3"].alignment = right
    ws.merge_cells("D3:E3")
    ws["D3"] = "تفاصيل سريعة"
    ws["D3"].fill = header_fill
    ws["D3"].font = white_font
    ws["D3"].alignment = right

    ws["A7"] = "إجمالي رسوم الحوالات"
    ws["A7"].font = label_font
    ws["A7"].fill = soft
    ws["A7"].alignment = right
    ws["A7"].border = border
    ws["B7"] = SUMMARY["transfer_fees_total"]
    ws["B7"].number_format = '#,##0.00 "SAR"'
    ws["B7"].font = bold_font
    ws["B7"].fill = soft
    ws["B7"].alignment = center
    ws["B7"].border = border

    for label, value, cell_ref in summary_items:
        if cell_ref.startswith("B"):
            row = int(cell_ref[1:])
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = label_font
            ws[f"A{row}"].fill = soft
            ws[f"A{row}"].alignment = right
            ws[f"A{row}"].border = border
            ws[cell_ref] = value
            ws[cell_ref].font = body_font if not isinstance(value, (int, float)) else bold_font
            ws[cell_ref].fill = soft
            ws[cell_ref].alignment = right if isinstance(value, str) else center
            ws[cell_ref].border = border
        else:
            row = int(cell_ref[1:])
            ws[f"D{row}"] = label
            ws[f"D{row}"].font = label_font
            ws[f"D{row}"].fill = soft
            ws[f"D{row}"].alignment = right
            ws[f"D{row}"].border = border
            ws[cell_ref] = value
            ws[cell_ref].font = body_font if not isinstance(value, (int, float)) else bold_font
            ws[cell_ref].fill = soft
            ws[cell_ref].alignment = right if isinstance(value, str) else center
            ws[cell_ref].border = border

    headers_row = 9
    headers = ["التاريخ", "نوع المصروف", "المبلغ"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(headers_row, idx, header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    start_row = 10
    expense_rows = [row for row in ROWS if row["move_type"] == "مصروف"]
    for i, row in enumerate(expense_rows, start=start_row):
        fill = red
        ws.cell(i, 1, row["date"])
        ws.cell(i, 1).number_format = "yyyy/mm/dd"
        ws.cell(i, 2, row["expense_type"] or "-")
        ws.cell(i, 3, row["amount"])
        ws.cell(i, 3).number_format = '#,##0.00 "SAR"'
        for c in range(1, 4):
            ws.cell(i, c).fill = fill
            ws.cell(i, c).font = body_font
            ws.cell(i, c).alignment = wrap if c == 2 else center
            ws.cell(i, c).border = border

    ws.freeze_panes = "A10"
    ws.auto_filter.ref = f"A9:C{start_row + len(expense_rows) - 1}"

    widths = {"A": 14, "B": 24, "C": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[9].height = 24
    for row in range(10, 10 + len(expense_rows)):
        ws.row_dimensions[row].height = 32

    # Place after "العمليات المالية" for consistent ordering.
    idx = wb.sheetnames.index("العمليات المالية") + 1
    wb._sheets.remove(ws)
    wb._sheets.insert(idx, ws)

    wb.save(SOURCE)


if __name__ == "__main__":
    build_statement_sheet()
