#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from typing import Iterable


PROJECT_DIR = Path(__file__).resolve().parents[1]
PYTHON = sys.executable
SOURCE_DIR = PROJECT_DIR / "source"
CANONICAL_SOURCE = SOURCE_DIR / "lead6.xlsx"
REPORT_PATH = PROJECT_DIR / "output" / "lead6_report.xlsx"
WEB_DIR = PROJECT_DIR / "web"
HOST_DIR = Path.home() / "lead6_host"
STATE_DIR = Path.home() / ".cache" / "lead6"
STATE_PATH = STATE_DIR / "auto_update_state.json"
LOCK_PATH = STATE_DIR / "auto_update.lock"
LOG_PATH = Path.home() / "Library" / "Logs" / "lead6-auto-update.log"

DEFAULT_SOURCE_CANDIDATES = [
    Path.home() / "Library" / "CloudStorage" / "GoogleDrive-gf.smartas@gmail.com" / "My Drive" / "lead6.xlsx",
    PROJECT_DIR / "source" / "lead6.xlsx",
    PROJECT_DIR / "البيانات الخام" / "بيانات .xlsx",
]

TRACKED_OUTPUTS = [
    PROJECT_DIR / "output" / "lead6_report.xlsx",
    PROJECT_DIR / "web" / "index.html",
    PROJECT_DIR / "web" / "gf_logo_current_clean.png",
    PROJECT_DIR / "web" / "gf_logo_transparent.png",
    PROJECT_DIR / "web" / "gf_logo_transparent_2x.png",
    PROJECT_DIR / "scripts" / "auto_update.py",
    PROJECT_DIR / "README.md",
    PROJECT_DIR / "launchd" / "com.ghaida.lead6.autoupdate.plist",
]


def log(message: str) -> None:
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with LOG_PATH.open("a", encoding="utf-8") as fh:
        fh.write(line + "\n")


def run(cmd: list[str], cwd: Path = PROJECT_DIR) -> None:
    log(f"RUN {' '.join(cmd)}")
    env = os.environ.copy()
    env["PATH"] = os.pathsep.join([
        str(Path.home() / "bin"),
        "/Library/Developer/CommandLineTools/usr/bin",
        "/usr/local/bin",
        "/opt/homebrew/bin",
        "/usr/bin",
        "/bin",
        "/usr/sbin",
        "/sbin",
    ])
    completed = subprocess.run(
        cmd,
        cwd=str(cwd),
        env=env,
        text=True,
        capture_output=True,
    )
    if completed.stdout:
        log(completed.stdout.strip())
    if completed.stderr:
        log(completed.stderr.strip())
    if completed.returncode != 0:
        raise RuntimeError(f"Command failed with exit code {completed.returncode}: {' '.join(cmd)}")


def ensure_dirs() -> None:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    HOST_DIR.mkdir(parents=True, exist_ok=True)
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)


def candidate_paths() -> list[Path]:
    raw = os.environ.get("LEAD6_SOURCE_PATHS", "").strip()
    if raw:
        candidates = [Path(item).expanduser() for item in raw.split(os.pathsep) if item.strip()]
        return candidates
    return DEFAULT_SOURCE_CANDIDATES


def file_signature(path: Path) -> str:
    stat = path.stat()
    return f"{stat.st_mtime_ns}:{stat.st_size}"


def is_valid_raw_workbook(path: Path) -> bool:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = set(wb.sheetnames)
        shipment_sheet = "Sheet1" if "Sheet1" in sheet_names else "الشحنات" if "الشحنات" in sheet_names else None
        ops_sheet = "Sheet2" if "Sheet2" in sheet_names else "العمليات المالية" if "العمليات المالية" in sheet_names else None
        if shipment_sheet is None or ops_sheet is None:
            return False
        shipment_ws = wb[shipment_sheet]
        ops_ws = wb[ops_sheet]
        return shipment_ws.max_row > 1 and shipment_ws.max_column >= 10 and ops_ws.max_row > 1 and ops_ws.max_column >= 5
    except Exception:
        return False


def latest_source() -> Path | None:
    existing = [path for path in candidate_paths() if path.exists() and path.is_file() and is_valid_raw_workbook(path)]
    if not existing:
        return None
    return max(existing, key=lambda p: (p.stat().st_mtime_ns, p.stat().st_size, str(p)))


def load_state() -> dict[str, str]:
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_state(state: dict[str, str]) -> None:
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def copy_source(src: Path) -> None:
    if src.resolve() != CANONICAL_SOURCE.resolve():
        shutil.copy2(src, CANONICAL_SOURCE)
        log(f"Copied source: {src} -> {CANONICAL_SOURCE}")


def sync_host_assets() -> None:
    HOST_DIR.mkdir(parents=True, exist_ok=True)
    assets = [
        REPORT_PATH,
        WEB_DIR / "index.html",
        WEB_DIR / "gf_logo_current_clean.png",
        WEB_DIR / "gf_logo_transparent.png",
        WEB_DIR / "gf_logo_transparent_2x.png",
    ]
    for src in assets:
        if src.exists():
            shutil.copy2(src, HOST_DIR / src.name)
            log(f"Synced asset: {src.name}")


def build_outputs() -> None:
    run([PYTHON, str(PROJECT_DIR / "scripts" / "build_report.py")])
    run([PYTHON, str(PROJECT_DIR / "web" / "generate_site.py")])
    sync_host_assets()


def git_status_has_changes() -> bool:
    result = subprocess.run(
        ["git", "status", "--porcelain", "--",
         "output/lead6_report.xlsx",
         "web/index.html",
         "web/gf_logo_current_clean.png",
         "web/gf_logo_transparent.png",
         "web/gf_logo_transparent_2x.png",
         "scripts/auto_update.py",
         "README.md",
         "launchd/com.ghaida.lead6.autoupdate.plist"],
        cwd=str(PROJECT_DIR),
        text=True,
        capture_output=True,
        check=True,
    )
    return bool(result.stdout.strip())


def git_commit_push(message: str) -> None:
    run([
        "git", "add", "-f",
        "output/lead6_report.xlsx",
        "web/index.html",
        "web/gf_logo_current_clean.png",
        "web/gf_logo_transparent.png",
        "web/gf_logo_transparent_2x.png",
        "scripts/auto_update.py",
        "README.md",
        "launchd/com.ghaida.lead6.autoupdate.plist",
    ])
    diff = subprocess.run(
        ["git", "diff", "--cached", "--quiet"],
        cwd=str(PROJECT_DIR),
    )
    if diff.returncode == 0:
        log("No staged changes to commit.")
        return
    run(["git", "commit", "-m", message])
    run(["git", "push", "origin", "main"])


def process_once(force: bool = False) -> bool:
    ensure_dirs()
    source = latest_source()
    if source is None:
        log("No source workbook found.")
        return False

    signature = f"{source.resolve()}::{file_signature(source)}"
    state = load_state()
    if not force and state.get("signature") == signature:
        log("No source change detected.")
        return False

    copy_source(source)
    build_outputs()
    commit_message = f"Automated lead6 refresh {time.strftime('%Y-%m-%d %H:%M')}"
    if git_status_has_changes():
        git_commit_push(commit_message)
    else:
        log("Working tree already clean after rebuild.")

    save_state({
        "signature": signature,
        "source_path": str(source),
        "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
    })
    return True


def acquire_lock() -> None:
    import fcntl

    ensure_dirs()
    lock_file = LOCK_PATH.open("w", encoding="utf-8")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError as exc:
        raise SystemExit("Another lead6 auto-update process is already running.") from exc
    lock_file.write(str(os.getpid()))
    lock_file.flush()
    globals()["_LOCK_FILE"] = lock_file


def watch(interval: int) -> None:
    log(f"Watching for lead6 updates every {interval}s")
    process_once(force=False)
    while True:
        try:
            process_once(force=False)
        except Exception as exc:
            log(f"Update failed: {exc}")
        time.sleep(interval)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Automate lead6 rebuilds from the raw workbook.")
    parser.add_argument("--once", action="store_true", help="Run a single rebuild cycle and exit.")
    parser.add_argument("--watch", action="store_true", help="Keep watching for source changes.")
    parser.add_argument("--interval", type=int, default=int(os.environ.get("LEAD6_POLL_SECONDS", "60")), help="Polling interval in seconds.")
    parser.add_argument("--force", action="store_true", help="Rebuild even if the source signature has not changed.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    acquire_lock()
    try:
        if args.watch:
            watch(args.interval)
        else:
            changed = process_once(force=args.force or args.once)
            return 0 if changed else 0
    except KeyboardInterrupt:
        log("Stopped by user.")
        return 130
    except Exception as exc:
        log(f"Fatal error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
