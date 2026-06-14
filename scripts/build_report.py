import copy
import datetime as dt
import math
import os
import re
import subprocess
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parents[1]
INPUT = PROJECT_DIR / "source" / "lead6.xlsx"
REPORT_OUT = PROJECT_DIR / "output" / "lead6_report.xlsx"
DASHBOARD_OUT = PROJECT_DIR / "output" / "lead6_dashboard.xlsx"


def money(value):
    if value in (None, "", "-"):
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group()) if match else 0.0


def normalize_text(value):
    return str(value).strip() if value is not None else ""


def parse_ship_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        match = re.search(r"(\d{4})[/-](\d{2})[/-](\d{2})", value)
        if match:
            return dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def read_prices(ws):
    headers = {normalize_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    carrier_name = {
        "ارامكس - ARAMEX": "ارامكس",
        "aramex ( استلام من الفرع )": "ارامكس استلام",
        "SMSA - سمسا": "سمسا",
        "SMSA ( استلام من الفرع )": "سمسا استلام",
        "RedBox - ريدبوكس": "ريد بوكس",
        "JT Express": "JT Express",
    }

    extra_customer_gross = money(ws.cell(2, headers["السعر لكل كيلو  زيادة"]).value)
    extra_customer_net = money(ws.cell(3, headers["السعر لكل كيلو  زيادة"]).value) or extra_customer_gross * 0.85
    extra_platform_gross = money(ws.cell(4, headers["السعر لكل كيلو  زيادة"]).value)
    extra_platform_net = money(ws.cell(5, headers["السعر لكل كيلو  زيادة"]).value) or extra_platform_gross * 0.85
    extra_profit_net = extra_customer_net - extra_platform_net

    cod_customer_gross = money(ws.cell(2, headers["سعر توصيل cod"]).value)
    cod_platform_gross = money(ws.cell(4, headers["سعر توصيل cod"]).value)
    cod_profit_net = cod_customer_gross - cod_platform_gross

    return {
        "headers": headers,
        "carrier_name": carrier_name,
        "extra_customer_gross": extra_customer_gross,
        "extra_customer_net": extra_customer_net,
        "extra_platform_gross": extra_platform_gross,
        "extra_platform_net": extra_platform_net,
        "extra_profit_net": extra_profit_net,
        "cod_customer_gross": cod_customer_gross,
        "cod_platform_gross": cod_platform_gross,
        "cod_profit_net": cod_profit_net,
    }


def platform_base_cost(prices_ws, prices, merchant, carrier):
    headers = prices["headers"]
    carrier_key = prices["carrier_name"].get(carrier)
    special_merchants = {"عبدالرحمن المطيري", "مؤسسة اواني القيصرية التجارية"}

    if merchant in special_merchants and merchant in headers:
        col = headers[merchant]
        platform_gross = money(prices_ws.cell(4, col).value)
        platform_net = money(prices_ws.cell(5, col).value)
        value = platform_gross if platform_gross else platform_net
        return money(value), merchant, platform_net, platform_gross

    if carrier_key and carrier_key in headers:
        col = headers[carrier_key]
        platform_net = money(prices_ws.cell(5, col).value)
        platform_gross = money(prices_ws.cell(4, col).value)
        value = platform_gross if merchant in special_merchants else platform_net
        if value in (None, 0.0):
            value = platform_net if merchant in special_merchants else platform_gross
        return money(value), carrier_key, platform_net, platform_gross

    return 0.0, "غير موجود في شيت الأسعار", 0.0, 0.0


def customer_price_from_sheet(prices_ws, prices, merchant, carrier, fallback_charge=0.0):
    headers = prices["headers"]
    carrier_key = prices["carrier_name"].get(carrier)
    special_merchants = {"عبدالرحمن المطيري", "مؤسسة اواني القيصرية التجارية"}
    if merchant in headers:
        col = headers[merchant]
        gross = money(prices_ws.cell(2, col).value)
        net = gross if merchant in special_merchants else (money(prices_ws.cell(3, col).value) or gross * 0.85)
        if gross:
            return gross, net, merchant
    if carrier_key and carrier_key in headers:
        col = headers[carrier_key]
        gross = money(prices_ws.cell(2, col).value)
        net = money(prices_ws.cell(3, col).value) or gross * 0.85
        if gross:
            return gross, net, carrier_key
    if fallback_charge:
        gross = money(fallback_charge)
        return gross, gross * 0.85, "شيت الشحنات"
    return 0.0, 0.0, "غير موجود في شيت الأسعار"


def better_row(existing, candidate):
    if existing is None:
        return candidate
    _, existing_values = existing
    _, candidate_values = candidate
    existing_score = int(existing_values.get("status") not in ("", None)) + int(existing_values.get("date") is not None)
    candidate_score = int(candidate_values.get("status") not in ("", None)) + int(candidate_values.get("date") is not None)
    return candidate if candidate_score >= existing_score else existing


def collect_shipments(wb):
    ws = wb["الشحنات"]
    prices_ws = wb["الاسعار"]
    prices = read_prices(prices_ws)
    by_order = {}

    for row in range(2, ws.max_row + 1):
        order_id = ws.cell(row, 1).value
        if order_id in (None, ""):
            continue

        status = normalize_text(ws.cell(row, 13).value)
        ship_date = parse_ship_date(ws.cell(row, 14).value)
        if ship_date is None or ship_date.month != 6 or not (1 <= ship_date.day <= 13):
            continue

        merchant = normalize_text(ws.cell(row, 3).value)
        carrier = normalize_text(ws.cell(row, 11).value)
        payment = normalize_text(ws.cell(row, 10).value)
        shipping_charge = money(ws.cell(row, 7).value)
        weight = money(ws.cell(row, 12).value)
        cod = payment == "💵 COD"
        extra_kg = max(weight - 15, 0)
        included = status not in ("ملغي", "مسودة")

        customer_gross, customer_net, customer_source = customer_price_from_sheet(prices_ws, prices, merchant, carrier, shipping_charge)
        base_cost, price_source, platform_net, platform_gross = platform_base_cost(prices_ws, prices, merchant, carrier)
        base_profit = customer_net - base_cost if included and customer_net and base_cost else 0.0
        extra_profit = extra_kg * prices["extra_profit_net"] if included else 0.0
        cod_profit = prices["cod_profit_net"] if included and cod else 0.0
        total_profit = base_profit + extra_profit + cod_profit

        values = {
            "order_id": order_id,
            "tracking": ws.cell(row, 2).value,
            "merchant": merchant,
            "store": ws.cell(row, 4).value,
            "customer": ws.cell(row, 5).value,
            "city": ws.cell(row, 6).value,
            "shipping_charge": money(ws.cell(row, 7).value),
            "cod_amount": money(ws.cell(row, 8).value),
            "order_amount": money(ws.cell(row, 9).value),
            "payment": payment,
            "carrier": carrier,
            "weight": weight,
            "status": status,
            "date": ship_date,
            "included": "نعم" if included else "لا",
            "price_source": f"{customer_source} / {price_source}",
            "customer_gross": customer_gross,
            "customer_net": customer_net,
            "platform_net": platform_net,
            "platform_gross": platform_gross,
            "base_cost": base_cost,
            "extra_kg": extra_kg,
            "base_profit": base_profit,
            "extra_profit": extra_profit,
            "cod_profit": cod_profit,
            "total_profit": total_profit,
            "source_row": row,
        }
        by_order[order_id] = better_row(by_order.get(order_id), (row, values))

    shipments = [item[1] for item in by_order.values()]
    shipments.sort(key=lambda x: (x["date"] or dt.date(1900, 1, 1), int(x["order_id"]) if isinstance(x["order_id"], int) else 0))
    return shipments


def reset_sheet(wb, title, index=None):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, index if index is not None else len(wb.sheetnames))
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    return ws


def style_range(ws, min_row, max_row, min_col, max_col, fill=None, font=None, border=True):
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_widths(ws, widths):
    if isinstance(widths, dict):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        return
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_details(wb, shipments):
    ws = reset_sheet(wb, "تفاصيل شهر 6", 1)
    headers = [
        "رقم الطلب", "رقم التتبع", "التاجر", "المتجر", "العميل", "المدينة",
        "قيمة الشحن على العميل", "مبلغ COD", "قيمة الطلب", "الدفع", "شركة الشحن",
        "الوزن", "الحالة", "التاريخ", "داخل حساب الربح", "مصدر السعر",
        "سعر العميل الصافي", "تكلفة المنصة بدون ضريبة",
        "كيلوات زيادة", "ربح الشحنة", "ربح الوزن الزائد",
        "ربح COD", "إجمالي الربح", "صف المصدر"
    ]
    ws.append(headers)
    for shipment in shipments:
        ws.append([
            shipment["order_id"], shipment["tracking"], shipment["merchant"], shipment["store"],
            shipment["customer"], shipment["city"], shipment["shipping_charge"], shipment["cod_amount"],
            shipment["order_amount"], shipment["payment"], shipment["carrier"], shipment["weight"],
            shipment["status"], shipment["date"], shipment["included"], shipment["price_source"],
            shipment["customer_net"], shipment["base_cost"], shipment["extra_kg"], shipment["base_profit"],
            shipment["extra_profit"], shipment["cod_profit"], shipment["total_profit"], shipment["source_row"],
        ])

    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    style_range(ws, 1, 1, 1, len(headers), header_fill, header_font)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    widths = [11, 18, 24, 28, 20, 18, 16, 12, 13, 12, 20, 10, 13, 12, 14, 20, 15, 15, 12, 13, 15, 12, 13, 10]
    set_widths(ws, widths)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col in (7, 8, 9, 17, 18, 20, 21, 22, 23):
            row[col - 1].number_format = '#,##0.00 "ريال"'
        row[13].number_format = "yyyy-mm-dd"
        row[11].number_format = '0.00 "كجم"'
        row[18].number_format = '0'
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style_range(ws, 2, ws.max_row, 1, len(headers), border=True)

    total_row = ws.max_row + 2
    ws.cell(total_row, 16).value = "الإجمالي"
    ws.cell(total_row, 17).value = sum(shipment["customer_net"] for shipment in shipments)
    ws.cell(total_row, 18).value = sum(shipment["base_cost"] for shipment in shipments)
    ws.cell(total_row, 21).value = sum(shipment["extra_profit"] for shipment in shipments)
    ws.cell(total_row, 20).value = sum(shipment["base_profit"] for shipment in shipments)
    ws.cell(total_row, 23).value = sum(shipment["total_profit"] for shipment in shipments)
    for col in (16, 17, 18, 20, 21, 23):
        ws.cell(total_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in (17, 18, 20, 21, 23):
        ws.cell(total_row, col).number_format = '#,##0.00 "ريال"'
    style_range(ws, total_row, total_row, 16, 23, PatternFill("solid", fgColor="16324F"), Font(color="FFFFFF", bold=True))
    return ws


def aggregate(shipments):
    included = [s for s in shipments if s["included"] == "نعم"]
    excluded = [s for s in shipments if s["included"] != "نعم"]

    by_merchant = defaultdict(lambda: {"count": 0, "base": 0.0, "extra": 0.0, "cod": 0.0, "total": 0.0})
    by_carrier = defaultdict(lambda: {"count": 0, "base": 0.0, "extra": 0.0, "cod": 0.0, "total": 0.0})
    by_status = defaultdict(int)
    for s in shipments:
        by_status[s["status"] or "بدون حالة"] += 1
        if s["included"] != "نعم":
            continue
        for bucket, key in ((by_merchant, s["merchant"]), (by_carrier, s["carrier"])):
            bucket[key]["count"] += 1
            bucket[key]["base"] += s["base_profit"]
            bucket[key]["extra"] += s["extra_profit"]
            bucket[key]["cod"] += s["cod_profit"]
            bucket[key]["total"] += s["total_profit"]

    totals = {
        "unique": len(shipments),
        "included": len(included),
        "excluded": len(excluded),
        "base": sum(s["base_profit"] for s in included),
        "extra": sum(s["extra_profit"] for s in included),
        "cod": sum(s["cod_profit"] for s in included),
        "total": sum(s["total_profit"] for s in included),
        "cod_count": sum(1 for s in included if s["payment"] == "💵 COD"),
        "overweight_count": sum(1 for s in included if s["extra_kg"] > 0),
    }
    return totals, by_merchant, by_carrier, by_status


def write_summary(wb, shipments, finance_summary=None):
    ws = reset_sheet(wb, "ملخصات", 2)
    totals, by_merchant, by_carrier, by_status = aggregate(shipments)

    start_date = dt.date(2026, 6, 1)
    end_date = dt.date(2026, 6, 13)
    ws["A1"] = f"{start_date.isoformat()} - {end_date.isoformat()}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="16324F")
    ws.merge_cells("A1:H1")

    ws.append([])
    ws.append(["المؤشر", "القيمة"])
    summary_rows = [
        ("الشحنات الفريدة بعد إزالة التكرار", totals["unique"]),
        ("الشحنات الداخلة في حساب الربح", totals["included"]),
        ("الشحنات المستبعدة: ملغي/مسودة", totals["excluded"]),
        ("ربح الشحنات", totals["base"]),
        ("ربح الشحن الزائد", totals["extra"]),
        ("ربح COD", totals["cod"]),
        ("إيداعات البنك", finance_summary["bank"]["total"] if finance_summary else 0),
        ("إيداعات ميسر", finance_summary["moyasar"]["total"] if finance_summary else 0),
        ("صافي الخصومات والاستردادات", finance_summary["other_net"] if finance_summary else 0),
        ("إجمالي الإيداعات", (finance_summary["bank"]["total"] + finance_summary["moyasar"]["total"]) if finance_summary else 0),
        ("إجمالي الربح", totals["total"]),
        ("عدد شحنات COD", totals["cod_count"]),
        ("عدد الشحنات ذات وزن زائد", totals["overweight_count"]),
    ]
    for row in summary_rows:
        ws.append(row)

    style_range(ws, 3, ws.max_row, 1, 2, PatternFill("solid", fgColor="16324F"), Font(color="FFFFFF", bold=True))

    start = ws.max_row + 3
    ws.cell(start, 1).value = "ملخص حسب التاجر"
    ws.cell(start, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(start, 1).fill = PatternFill("solid", fgColor="16324F")
    headers = ["التاجر", "عدد الشحنات", "ربح الشحنات", "ربح الزائد", "ربح COD", "إجمالي الربح"]
    for col, header in enumerate(headers, 1):
        ws.cell(start + 1, col).value = header
    for name, data in sorted(by_merchant.items(), key=lambda x: x[1]["total"], reverse=True):
        ws.append([name, data["count"], data["base"], data["extra"], data["cod"], data["total"]])

    carrier_start = ws.max_row + 3
    ws.cell(carrier_start, 1).value = "ملخص حسب شركة الشحن"
    ws.cell(carrier_start, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(carrier_start, 1).fill = PatternFill("solid", fgColor="16324F")
    for col, header in enumerate(["شركة الشحن", "عدد الشحنات", "ربح الشحنات", "ربح الزائد", "ربح COD", "إجمالي الربح"], 1):
        ws.cell(carrier_start + 1, col).value = header
    for name, data in sorted(by_carrier.items(), key=lambda x: x[1]["total"], reverse=True):
        ws.append([name, data["count"], data["base"], data["extra"], data["cod"], data["total"]])

    status_start = ws.max_row + 3
    ws.cell(status_start, 1).value = "عدد الشحنات حسب الحالة"
    ws.cell(status_start, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(status_start, 1).fill = PatternFill("solid", fgColor="16324F")
    ws.cell(status_start + 1, 1).value = "الحالة"
    ws.cell(status_start + 1, 2).value = "العدد"
    for name, count in sorted(by_status.items(), key=lambda x: x[1], reverse=True):
        ws.append([name, count])

    set_widths(ws, [32, 15, 15, 15, 13, 15, 4, 4])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00 "ريال"'
    for row in (3, start + 1, carrier_start + 1, status_start + 1):
        style_range(ws, row, row, 1, 6 if row != status_start + 1 else 2, PatternFill("solid", fgColor="DDEBF7"), Font(bold=True))
    return ws, totals, by_merchant, by_carrier, by_status


def write_dashboard(wb, totals, by_merchant, by_carrier, shipments):
    ws = reset_sheet(wb, "Dashboard", 0)
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "D7"

    set_widths(
        ws,
        {
            "A": 10, "B": 12, "C": 12, "D": 16, "E": 16, "F": 16,
            "G": 16, "H": 16, "I": 16, "J": 16,
        },
    )
    for row in range(1, 60):
        ws.row_dimensions[row].height = 22

    # Sidebar
    for r in range(1, 60):
        for c in range(1, 4):
            ws.cell(r, c).fill = PatternFill("solid", fgColor="17202B")
            ws.cell(r, c).border = Border()
    ws.merge_cells("A1:C2")
    ws["A1"] = "L"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F9D8A")

    ws.merge_cells("A3:C4")
    ws["A3"] = "lead"
    ws["A3"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A5:C5")
    ws["A5"] = "Operations OS"
    ws["A5"].font = Font(color="AEB9C5", bold=True)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

    nav_items = [("⌘", "الرئيسية"), ("◫", "الشحنات"), ("◌", "المصاريف"), ("◍", "التقارير")]
    for idx, (icon, label) in enumerate(nav_items, start=7):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=3)
        ws.cell(idx, 1).value = f"{icon}  {label}"
        ws.cell(idx, 1).font = Font(color="C6D0DC", bold=True)
        ws.cell(idx, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor="243140" if label == "الرئيسية" else "17202B")

    ws.merge_cells("A13:C17")
    ws["A13"] = "SLA\n98.7%\nاستقرار العمليات خلال آخر 30 يوم"
    ws["A13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A13"].font = Font(color="FFFFFF", bold=True)
    ws["A13"].fill = PatternFill("solid", fgColor="202D39")

    # Main topbar
    ws.merge_cells("D1:J2")
    ws["D1"] = "لوحة قيادة أرباح lead"
    ws["D1"].font = Font(size=20, bold=True, color="18222D")
    ws["D1"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("D3:F4")
    ws["D3"] = "ابحث في بيانات lead"
    ws["D3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["D3"].font = Font(color="637083", italic=True)
    ws["D3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["D3"].border = Border(left=Side(style="thin", color="DCE3EA"), right=Side(style="thin", color="DCE3EA"), top=Side(style="thin", color="DCE3EA"), bottom=Side(style="thin", color="DCE3EA"))

    ws.merge_cells("G3:H4")
    ws["G3"] = "↻"
    ws["G3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G3"].font = Font(size=16, bold=True, color="18222D")
    ws["G3"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["G3"].border = Border(left=Side(style="thin", color="DCE3EA"), right=Side(style="thin", color="DCE3EA"), top=Side(style="thin", color="DCE3EA"), bottom=Side(style="thin", color="DCE3EA"))

    ws.merge_cells("I3:J4")
    ws["I3"] = "افتح التفاصيل"
    ws["I3"].hyperlink = str(REPORT_OUT)
    ws["I3"].style = "Hyperlink"
    ws["I3"].fill = PatternFill("solid", fgColor="17202B")
    ws["I3"].font = Font(color="FFFFFF", bold=True)
    ws["I3"].alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards
    cards = [
        ("إجمالي الربح", totals["total"], "17202B"),
        ("ربح الشحنات", totals["base"], "1F9D8A"),
        ("ربح الوزن الزائد", totals["extra"], "D89B2B"),
        ("ربح COD", totals["cod"], "2E9F61"),
        ("الشحنات المحسوبة", totals["included"], "243140"),
        ("المستبعدة", totals["excluded"], "E0644F"),
    ]
    positions = ["D7:E9", "F7:G9", "H7:I9", "J7:J9", "D10:F12", "G10:J12"]
    for (label, value, color), rng in zip(cards, positions):
        start = rng.split(":")[0]
        ws.merge_cells(rng)
        ws[start] = f"{label}\n{value:,.2f} ريال" if isinstance(value, float) else f"{label}\n{value}"
        ws[start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[start].font = Font(color="FFFFFF", bold=True, size=12)
        ws[start].fill = PatternFill("solid", fgColor=color)

    # Panels
    ws.merge_cells("D14:G15")
    ws["D14"] = "مؤشرات لوحة التحكم"
    ws["D14"].font = Font(color="FFFFFF", bold=True)
    ws["D14"].fill = PatternFill("solid", fgColor="17202B")
    ws["D14"].alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("H14:J15")
    ws["H14"] = "قناة النمو"
    ws["H14"].font = Font(color="FFFFFF", bold=True)
    ws["H14"].fill = PatternFill("solid", fgColor="17202B")
    ws["H14"].alignment = Alignment(horizontal="right", vertical="center")

    merchant_rows = sorted(by_merchant.items(), key=lambda x: x[1]["total"], reverse=True)[:5]
    ws["D16"] = "التاجر"
    ws["E16"] = "إجمالي الربح"
    ws["D16"].font = ws["E16"].font = Font(bold=True, color="17202B")
    ws["D16"].fill = ws["E16"].fill = PatternFill("solid", fgColor="DDEBF7")
    for idx, (name, data) in enumerate(merchant_rows, start=17):
        ws.cell(idx, 4).value = name
        ws.cell(idx, 5).value = data["total"]
        ws.cell(idx, 5).number_format = '#,##0.00 "ريال"'
    style_range(ws, 16, 16 + len(merchant_rows), 4, 5)

    ws["H16"] = "شركاء الشحن"
    ws["I16"] = "النسبة"
    ws["J16"] = "المؤشر"
    for c in ("H16", "I16", "J16"):
        ws[c].font = Font(bold=True, color="17202B")
        ws[c].fill = PatternFill("solid", fgColor="DDEBF7")
    partner_rows = [("SMSA", 74), ("ARAMEX", 46), ("RED BOX", 18)]
    for idx, (name, pct) in enumerate(partner_rows, start=17):
        ws.cell(idx, 8).value = name
        ws.cell(idx, 9).value = f"{pct}%"
        ws.cell(idx, 10).value = ""
        ws.cell(idx, 10).fill = PatternFill("solid", fgColor="1F9D8A" if pct > 60 else "D89B2B" if pct > 30 else "E0644F")
    style_range(ws, 16, 19, 8, 10)

    # Charts
    bar = BarChart()
    bar.title = "أعلى التجار حسب الربح"
    bar.y_axis.title = "ريال"
    bar.x_axis.title = "التاجر"
    data = Reference(ws, min_col=5, min_row=16, max_row=16 + len(merchant_rows))
    cats = Reference(ws, min_col=4, min_row=17, max_row=16 + len(merchant_rows))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 7
    bar.width = 10
    ws.add_chart(bar, "D21")

    pie = PieChart()
    pie.title = "توزيع الربح"
    ws["H21"] = "النوع"
    ws["I21"] = "القيمة"
    breakdown = [("ربح الشحنات", totals["base"]), ("ربح الوزن الزائد", totals["extra"]), ("ربح COD", totals["cod"])]
    for i, (label, val) in enumerate(breakdown, start=22):
        ws.cell(i, 8).value = label
        ws.cell(i, 9).value = val
        ws.cell(i, 9).number_format = '#,##0.00 "ريال"'
    pdata = Reference(ws, min_col=9, min_row=21, max_row=24)
    pcats = Reference(ws, min_col=8, min_row=22, max_row=24)
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(pcats)
    pie.height = 7
    pie.width = 8.5
    ws.add_chart(pie, "H25")

    # Latest shipments table
    ws.merge_cells("D40:J41")
    ws["D40"] = "آخر الطلبات من lead"
    ws["D40"].font = Font(color="FFFFFF", bold=True)
    ws["D40"].fill = PatternFill("solid", fgColor="17202B")
    headers = ["رقم التتبع", "العميل/التاجر", "المدينة", "الحالة"]
    for c, h in enumerate(headers, start=4):
        ws.cell(42, c).value = h
        ws.cell(42, c).font = Font(bold=True, color="17202B")
        ws.cell(42, c).fill = PatternFill("solid", fgColor="DDEBF7")
    latest = [r for r in shipments if r["included"] == "نعم"][:3]
    for row_idx, rec in enumerate(latest, start=43):
        ws.cell(row_idx, 4).value = rec["tracking"]
        ws.cell(row_idx, 5).value = rec["merchant"]
        ws.cell(row_idx, 6).value = rec["city"]
        ws.cell(row_idx, 7).value = rec["status"]
    style_range(ws, 42, 42 + len(latest), 4, 7)

    ws.merge_cells("D50:J52")
    ws["D50"] = "ملاحظة: ملف التفاصيل منفصل، والزر بالأعلى يفتح عليه مباشرة."
    ws["D50"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws["D50"].fill = PatternFill("solid", fgColor="FFF2CC")
    style_range(ws, 50, 52, 4, 10)


def style_prices_sheet(wb):
    if "الاسعار" not in wb.sheetnames:
        return
    ws = wb["الاسعار"]
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"
    set_widths(ws, {"A": 22, "B": 14, "C": 16, "D": 14, "E": 16, "F": 14, "G": 18, "H": 18, "I": 14, "J": 16, "K": 18})
    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_fills = {
        2: "EAF4F4",
        3: "F7FBFB",
        4: "FFF1E8",
        5: "FFF8F0",
    }
    for r, color in row_fills.items():
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="D9E2EC"),
                right=Side(style="thin", color="D9E2EC"),
                top=Side(style="thin", color="D9E2EC"),
                bottom=Side(style="thin", color="D9E2EC"),
            )
            if c > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "ريال"'
    for r in range(1, 6):
        ws.row_dimensions[r].height = 24
    style_range(ws, 1, 5, 1, ws.max_column)


def style_operations_sheet(wb):
    if "العمليات المالية" not in wb.sheetnames:
        return
    ws = wb["العمليات المالية"]
    raw_rows = []
    summary = {
        "bank": {"count": 0, "total": 0.0},
        "moyasar": {"count": 0, "total": 0.0},
        "shipping_return": {"count": 0, "total": 0.0},
        "tax_deduction": {"count": 0, "total": 0.0},
        "shipping_refund": {"count": 0, "total": 0.0},
        "other_net": 0.0,
        "other": {"count": 0, "total": 0.0},
    }
    for r in range(2, ws.max_row + 1):
        op_id = ws.cell(r, 1).value
        op_date = ws.cell(r, 2).value
        name = normalize_text(ws.cell(r, 3).value)
        kind = normalize_text(ws.cell(r, 4).value)
        amount = money(ws.cell(r, 5).value)
        note = normalize_text(ws.cell(r, 6).value)
        ref = ws.cell(r, 7).value
        if not any(v not in (None, "") for v in (op_id, op_date, name, kind, amount, note, ref)):
            continue
        summary["other_net"] += amount
        channel = "البنك" if "تحويل بنكي" in note else "ميسر" if "Moyasar" in note else "أخرى"
        if kind == "إيداع":
            if channel == "البنك":
                summary["bank"]["count"] += 1
                summary["bank"]["total"] += amount
            elif channel == "ميسر":
                summary["moyasar"]["count"] += 1
                summary["moyasar"]["total"] += amount
        elif kind == "admin_deduction" and "خصم تكلفة شحنة مرتجعة" in note:
            summary["shipping_return"]["count"] += 1
            summary["shipping_return"]["total"] += abs(amount)
        elif kind == "admin_deduction" and "خصم ضريبة القيمة المضافة" in note:
            summary["tax_deduction"]["count"] += 1
            summary["tax_deduction"]["total"] += abs(amount)
        elif kind == "استرداد" and "استرداد تكلفة شحن" in note:
            summary["shipping_refund"]["count"] += 1
            summary["shipping_refund"]["total"] += amount
        else:
            summary["other"]["count"] += 1
            summary["other"]["total"] += amount
        raw_rows.append({
            "op_id": op_id,
            "date": op_date,
            "name": name,
            "kind": kind,
            "amount": amount,
            "note": note,
            "ref": ref,
            "channel": channel,
        })

    # Rebuild as a clean sheet with summary + detailed table.
    del wb["العمليات المالية"]
    ws = wb.create_sheet("العمليات المالية", 3)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"

    ws.merge_cells("A1:H1")
    ws["A1"] = "العمليات المالية"
    ws["A1"].fill = PatternFill("solid", fgColor="16324F")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "القناة"
    ws["B3"] = "عدد الإيداعات"
    ws["C3"] = "الإجمالي"
    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor="16324F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A4"] = "البنك"
    ws["B4"] = summary["bank"]["count"]
    ws["C4"] = summary["bank"]["total"]
    ws["A5"] = "ميسر"
    ws["B5"] = summary["moyasar"]["count"]
    ws["C5"] = summary["moyasar"]["total"]
    ws["A6"] = "خصم تكلفة شحنة مرتجعة"
    ws["B6"] = summary["shipping_return"]["count"]
    ws["C6"] = summary["shipping_return"]["total"]
    ws["A7"] = "خصم الضريبة"
    ws["B7"] = summary["tax_deduction"]["count"]
    ws["C7"] = summary["tax_deduction"]["total"]
    ws["A8"] = "استرداد تكلفة شحن"
    ws["B8"] = summary["shipping_refund"]["count"]
    ws["C8"] = summary["shipping_refund"]["total"]
    ws["A9"] = "المجموع"
    ws["B9"] = summary["bank"]["count"] + summary["moyasar"]["count"]
    ws["C9"] = summary["bank"]["total"] + summary["moyasar"]["total"]
    for r in range(4, 10):
        for c in range(1, 4):
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(r, c).border = Border(left=Side(style="thin", color="D9E2EC"), right=Side(style="thin", color="D9E2EC"), top=Side(style="thin", color="D9E2EC"), bottom=Side(style="thin", color="D9E2EC"))
            if c == 3:
                ws.cell(r, c).number_format = '#,##0.00 "ريال"'
    ws["A11"] = "رقم العملية"
    ws["B11"] = "التاريخ"
    ws["C11"] = "الاسم"
    ws["D11"] = "النوع"
    ws["E11"] = "المبلغ"
    ws["F11"] = "البيان"
    ws["G11"] = "المرجع"
    ws["H11"] = "القناة"
    for cell in ws[11]:
        cell.fill = PatternFill("solid", fgColor="16324F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sort detailed rows by date descending, then write them.
    raw_rows.sort(key=lambda x: (x["date"] or dt.datetime(1900, 1, 1), str(x["op_id"])), reverse=True)
    for idx, row in enumerate(raw_rows, start=12):
        ws.cell(idx, 1).value = row["op_id"]
        ws.cell(idx, 2).value = row["date"]
        ws.cell(idx, 3).value = row["name"]
        ws.cell(idx, 4).value = row["kind"]
        ws.cell(idx, 5).value = row["amount"]
        ws.cell(idx, 6).value = row["note"]
        ws.cell(idx, 7).value = row["ref"]
        ws.cell(idx, 8).value = row["channel"]
        for c in range(1, 9):
            ws.cell(idx, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(idx, c).border = Border(left=Side(style="thin", color="D9E2EC"), right=Side(style="thin", color="D9E2EC"), top=Side(style="thin", color="D9E2EC"), bottom=Side(style="thin", color="D9E2EC"))
        ws.cell(idx, 2).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(idx, 5).number_format = '#,##0.00 "ريال"'

    set_widths(ws, {"A": 14, "B": 20, "C": 24, "D": 18, "E": 14, "F": 46, "G": 18, "H": 12})
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24
    return summary


def main():
    wb = load_workbook(INPUT)
    style_prices_sheet(wb)
    style_operations_sheet(wb)
    shipments = collect_shipments(wb)
    report = load_workbook(INPUT)
    style_prices_sheet(report)
    finance_summary = style_operations_sheet(report)
    write_details(report, shipments)
    _, totals, by_merchant, by_carrier, _ = write_summary(report, shipments, finance_summary)
    preferred = ["ملخصات", "تفاصيل شهر 6", "الاسعار", "العمليات المالية"]
    report._sheets = [report[name] for name in preferred if name in report.sheetnames]
    report.active = report.sheetnames.index("ملخصات")
    report.save(REPORT_OUT)

    subprocess.run([sys.executable, str(PROJECT_DIR / "scripts" / "update_statement.py")], check=True)

    dashboard = load_workbook(REPORT_OUT)
    write_dashboard(dashboard, totals, by_merchant, by_carrier, shipments)
    dashboard._sheets = [dashboard["Dashboard"]]
    dashboard.active = 0
    dashboard.save(DASHBOARD_OUT)
    print(REPORT_OUT)
    print(DASHBOARD_OUT)
    print(totals)


if __name__ == "__main__":
    main()
