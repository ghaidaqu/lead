"""On-demand Excel export built in-memory from PostgreSQL.

No .xlsx file is ever written to disk — the workbook is streamed straight to the
client. `openpyxl` is used here and only here.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scripts.bank_statement import filtered_bank_transfer_fees, summarize_bank_statement

SHIPMENT_COLUMNS = [
    ("order_id", "رقم الطلب"),
    ("merchant_name", "التاجر"),
    ("store_name", "المتجر"),
    ("customer_name", "العميل"),
    ("city", "المدينة"),
    ("carrier", "شركة الشحن"),
    ("payment_type", "الدفع"),
    ("status", "الحالة"),
    ("shipment_date", "تاريخ الشحن"),
    ("weight", "الوزن"),
    ("cod_amount", "مبلغ COD"),
    ("customer_price_net", "صافي العميل"),
    ("platform_cost_net", "صافي المنصة"),
    ("base_profit", "ربح الشحنة"),
    ("extra_profit", "ربح الوزن الزائد"),
    ("cod_profit", "ربح COD"),
    ("total_profit", "إجمالي الربح"),
    ("actual_revenue", "الإيراد الفعلي"),
    ("actual_base_cost", "التكلفة الأساسية الفعلية"),
    ("actual_extra_cost", "رسوم وزن/COD الفعلية"),
    ("actual_profit", "صافي الربح الفعلي"),
    ("included_in_profit", "محتسب"),
]

WALLET_COLUMNS = [
    ("transaction_key", "Transaction Key"),
    ("transaction_date", "Transaction Date"),
    ("user_name", "User"),
    ("description", "Description"),
    ("amount", "Amount"),
    ("transaction_type", "Type"),
    ("balance_before", "Balance Before"),
    ("balance_after", "Balance After"),
    ("source_page", "Source Page"),
    ("raw_payload", "Raw Payload"),
]

PAYMENT_COLUMNS = [
    ("payment_key", "Payment Key"),
    ("payment_date", "Payment Date"),
    ("customer_name", "Customer"),
    ("amount", "Amount"),
    ("method", "Method"),
    ("status", "Status"),
    ("source_page", "Source Page"),
    ("raw_payload", "Raw Payload"),
]

_HEADER_FILL = PatternFill("solid", fgColor="7A2438")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TOTAL_FONT = Font(bold=True)
_TOTAL_COLUMNS = {"actual_revenue", "actual_base_cost", "actual_profit"}


def _style_header(ws) -> None:
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _append_totals_row(ws, columns: list[tuple[str, str]], totals: dict[str, float]) -> None:
    row = [""] * len(columns)
    row[0] = "الإجمالي"
    for idx, (column, _) in enumerate(columns):
        if column in totals:
            row[idx] = round(totals[column], 2)
    ws.append(row)

    total_row_idx = ws.max_row
    ws.row_dimensions[total_row_idx].height = 22
    for cell in ws[total_row_idx]:
        cell.font = _TOTAL_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_report_xlsx(conn, date_from=None, date_to=None) -> bytes:
    """Build the report workbook from the DB and return it as bytes."""
    shipment_where, shipment_params = _date_where("shipment_date", date_from, date_to)
    wallet_where, wallet_params = _date_where("transaction_date", date_from, date_to)
    payment_where, payment_params = _date_where("payment_date", date_from, date_to)
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT {', '.join(c for c, _ in SHIPMENT_COLUMNS)} "
            f"FROM shipments {shipment_where} ORDER BY shipment_date NULLS LAST, order_id",
            shipment_params,
        )
        rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            f"SELECT {', '.join(c for c, _ in WALLET_COLUMNS)} "
            f"FROM wallet_transactions {wallet_where} ORDER BY transaction_date NULLS LAST, transaction_key",
            wallet_params,
        )
        wallet_rows = [dict(r) for r in cur.fetchall()]
        cur.execute(
            f"SELECT {', '.join(c for c, _ in PAYMENT_COLUMNS)} "
            f"FROM payments {payment_where} ORDER BY payment_date NULLS LAST, payment_key",
            payment_params,
        )
        payment_rows = [dict(r) for r in cur.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "الشحنات"
    ws.sheet_view.rightToLeft = True
    ws.append([label for _, label in SHIPMENT_COLUMNS])
    totals = {column: 0.0 for column in _TOTAL_COLUMNS}
    for row in rows:
        ws.append([_fmt(row.get(col)) for col, _ in SHIPMENT_COLUMNS])
        for column in _TOTAL_COLUMNS:
            totals[column] += float(row.get(column) or 0)
    _append_totals_row(ws, SHIPMENT_COLUMNS, totals)
    _style_header(ws)

    summary = wb.create_sheet("الملخص")
    summary.sheet_view.rightToLeft = True
    summary.append(["البند", "القيمة"])
    summary.append(["الفترة من", date_from.isoformat() if date_from else ""])
    summary.append(["الفترة إلى", date_to.isoformat() if date_to else ""])
    summary.append(["عدد الشحنات", len(rows)])
    summary.append(["صافي الربح الفعلي", round(totals["actual_profit"], 2)])
    summary.append(["المحتسبة في الربح", sum(1 for r in rows if r.get("included_in_profit"))])
    bank_statement = summarize_bank_statement(date_from, date_to)
    summary.append(["رسوم الحوالات البنكية", bank_statement["summary"]["transfer_fees_total"]])
    _style_header(summary)

    _append_raw_sheet(wb, "Raw_Wallet", WALLET_COLUMNS, wallet_rows)
    _append_raw_sheet(wb, "Raw_Payments", PAYMENT_COLUMNS, payment_rows)
    _append_bank_fees_sheet(wb, filtered_bank_transfer_fees(date_from, date_to))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _date_where(column: str, date_from, date_to) -> tuple[str, list[Any]]:
    clauses = []
    params: list[Any] = []
    if date_from:
        clauses.append(f"{column} >= %s")
        params.append(date_from)
    if date_to:
        clauses.append(f"{column} < %s")
        import datetime as _dt
        params.append(date_to + _dt.timedelta(days=1))
    return (("WHERE " + " AND ".join(clauses)) if clauses else "", params)


def _append_raw_sheet(wb: Workbook, title: str, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append([label for _, label in columns])
    for row in rows:
        ws.append([_fmt(row.get(col)) for col, _ in columns])
    _style_header(ws)


def _append_bank_fees_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Bank_Transfer_Fees")
    ws.sheet_view.rightToLeft = True
    ws.append(["التاريخ", "نوع المصروف", "المبلغ", "المصدر"])
    for row in rows:
        ws.append([row["date"], row["expense_type"], row["amount"], row["source"]])
    _style_header(ws)


def _fmt(value: Any):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    if isinstance(value, bool):
        return "نعم" if value else "لا"
    if isinstance(value, (list, dict)):
        import json
        return json.dumps(value, ensure_ascii=False, default=str)
    try:
        return float(value)
    except (TypeError, ValueError):
        return str(value)
