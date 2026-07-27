from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook


DATE_KEYS = ("التاريخ", "تاريخ العملية", "date", "transaction date")
DESC_KEYS = ("البيان", "الوصف", "التفاصيل", "ملاحظات", "description", "details", "narrative")
DEBIT_KEYS = ("مدين", "مسحوب", "خصم", "debit", "withdrawal")
CREDIT_KEYS = ("دائن", "إيداع", "credit", "deposit")
AMOUNT_KEYS = ("المبلغ", "amount", "قيمة العملية")


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value or "")).strip()).lower()


def _pick(row: dict[str, Any], keys: tuple[str, ...]):
    for key, value in row.items():
        label = _norm(key)
        if any(k in label for k in keys):
            return value
    return None


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).replace(",", "").replace("ر.س", "").replace("sar", "").strip()
    text = text.replace("−", "-")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def _date(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return None


def classify(description: str, direction: str) -> str:
    text = _norm(description)
    rules = (
        ("رسوم حوالات", ("رسوم حوال", "transfer fee", "رسوم تحويل")),
        ("تحويل داخلي", ("عملية تحويل داخلية", "internal transfer")),
        ("تسويات بطاقات", ("مستحقات البطاقات", "card settlement")),
        ("رواتب", ("راتب", "رواتب", "salary", "payroll")),
        ("اشتراكات وخدمات", ("اشتراك", "subscription", "software", "خدمة")),
        ("مشتريات ومصاريف تشغيلية", ("شراء", "مدى", "pos", "مشتريات", "فاتورة")),
        ("تحويلات صادرة", ("صادرة", "تحويل صادر", "outgoing transfer")),
        ("إيداعات عملاء", ("إيداع", "واردة", "تحويل وارد", "incoming transfer")),
        ("ضريبة", ("ضريبة", "vat")),
    )
    for category, words in rules:
        if any(word in text for word in words):
            return category
    return "إيداع غير مصنف" if direction == "credit" else "مصروف غير مصنف"


def normalize_row(row: dict[str, Any], source_file: str) -> dict[str, Any] | None:
    day = _date(_pick(row, DATE_KEYS))
    description = str(_pick(row, DESC_KEYS) or "").strip()
    debit = _number(_pick(row, DEBIT_KEYS))
    credit = _number(_pick(row, CREDIT_KEYS))
    amount = _number(_pick(row, AMOUNT_KEYS))
    if credit not in (None, 0):
        direction, value = "credit", abs(credit)
    elif debit not in (None, 0):
        direction, value = "debit", abs(debit)
    elif amount not in (None, 0):
        direction, value = ("debit", abs(amount)) if amount < 0 else ("credit", amount)
    else:
        return None
    if not day or not description:
        return None
    fingerprint = hashlib.sha256(f"{day}|{description}|{value:.2f}|{direction}".encode("utf-8")).hexdigest()
    return {
        "transaction_date": day,
        "description": description,
        "amount": round(value, 2),
        "direction": direction,
        "category": classify(description, direction),
        "include": True,
        "fingerprint": fingerprint,
        "source_file": source_file,
    }


def _dict_rows(values: list[list[Any]]) -> list[dict[str, Any]]:
    header_index = next(
        (i for i, row in enumerate(values[:20]) if any(any(k in _norm(v) for k in DATE_KEYS) for v in row)),
        None,
    )
    if header_index is None:
        raise ValueError("bank_statement_headers_not_found")
    headers = [str(v or "").strip() for v in values[header_index]]
    return [dict(zip(headers, row)) for row in values[header_index + 1 :] if any(v not in (None, "") for v in row)]


def parse_bank_file(filename: str, data: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        values = [list(row) for row in workbook.active.iter_rows(values_only=True)]
        raw_rows = _dict_rows(values)
    elif lower.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        raw_rows = list(csv.DictReader(io.StringIO(text)))
    elif lower.endswith(".pdf"):
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ValueError("pdf_support_unavailable") from exc
        text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(data)).pages)
        raw_rows = []
        amounts = re.compile(
            r"^(?P<balance>-?\d[\d,]*\.\d{2})\s+sar\s+"
            r"(?P<credit>-?\d[\d,]*\.\d{2})\s+sar\s+"
            r"(?P<debit>-?\d[\d,]*\.\d{2})\s+sar\s+(?P<description>.+)$",
            re.IGNORECASE,
        )
        date_line = re.compile(r"^\d{4}[/-]\d{1,2}[/-]\d{1,2}$")
        pending = None
        for original in text.splitlines():
            line = _norm(original)
            match = amounts.match(line)
            if match:
                pending = {
                    "دائن": match["credit"],
                    "مدين": match["debit"],
                    "parts": [match["description"]],
                }
                continue
            if pending and date_line.match(line):
                description = " ".join(
                    part for part in pending.pop("parts")
                    if part not in {"الوقت"} and not part.isdigit()
                )
                raw_rows.append({"التاريخ": line, "الوصف": description, **pending})
                pending = None
                continue
            if pending and line and not line.startswith(("ref. no", "note ")):
                pending["parts"].append(line)
    else:
        raise ValueError("unsupported_file_type")
    rows = [normalize_row(row, filename) for row in raw_rows]
    return [row for row in rows if row]
