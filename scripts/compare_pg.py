from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from scripts.db_store import compare_counts, db_enabled, get_conn


SOURCE = PROJECT_DIR / "source" / "lead6.xlsx"


def workbook_counts() -> dict[str, int]:
    wb = load_workbook(SOURCE)
    counts = {}
    for name in ("Raw_Shipments", "Raw_Wallet", "Raw_Payments", "Raw_COD"):
        ws = wb[name]
        counts[name] = max(0, ws.max_row - 1)
    return counts


def main() -> int:
    if not db_enabled():
        print(json.dumps({"enabled": False}, ensure_ascii=False, indent=2))
        return 0
    with get_conn() as conn:
        pg = compare_counts(conn)
    wb = workbook_counts()
    report = {"enabled": True, "postgres": pg, "workbook": wb}
    report["diff"] = {
        "shipments": pg["shipments"] - wb["Raw_Shipments"],
        "wallet_transactions": pg["wallet_transactions"] - wb["Raw_Wallet"],
        "payments": pg["payments"] - wb["Raw_Payments"],
        "cod_collections": pg["cod_collections"] - wb["Raw_COD"],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
