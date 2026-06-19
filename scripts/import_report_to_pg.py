from __future__ import annotations

import datetime as dt
import json
import random
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from scripts.db_store import db_enabled, ensure_schema, finish_sync_run, get_conn, make_sync_run, upsert_rows


WORKBOOK = PROJECT_DIR / "output" / "lead6_report.xlsx"


def norm(v):
    return "" if v is None else str(v).strip()


def num(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def datev(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def shipment_payload(row):
    return {
        "order_id": norm(row[0]),
        "tracking_number": norm(row[1]),
        "merchant_name": norm(row[2]),
        "store_name": norm(row[3]),
        "customer_name": norm(row[4]),
        "city": norm(row[5]),
        "carrier": norm(row[10]),
        "payment_type": norm(row[9]),
        "status": norm(row[12]),
        "shipment_date": datev(row[13]),
        "delivery_date": None,
        "weight": num(row[11]),
        "cod_amount": num(row[7]),
        "shipping_charge": num(row[6]),
        "customer_price_gross": num(row[16]),
        "customer_price_net": num(row[16]),
        "platform_cost_gross": num(row[17]),
        "platform_cost_net": num(row[17]),
        "base_profit": num(row[19]),
        "extra_kg": num(row[18]),
        "extra_profit": num(row[20]),
        "cod_profit": num(row[21]),
        "total_profit": num(row[22]),
        "included_in_profit": norm(row[14]) == "نعم",
        "source_row": int(row[23]) if len(row) > 23 and str(row[23]).isdigit() else None,
        "source_hash": norm(row[23]),
        "raw_payload": row,
    }


def wallet_payload(row, category):
    key = f"{row[0]}|{row[1]}|{row[2]}|{row[3]}|{category}"
    amount = num(row[2])
    if category == "bank":
        amount = num(row[2])
    return {
        "transaction_key": key,
        "transaction_date": None,
        "user_name": norm(row[0]),
        "description": norm(row[0]),
        "amount": amount,
        "transaction_type": category,
        "balance_before": None,
        "balance_after": None,
        "source_page": "العمليات المالية",
        "raw_payload": row,
    }


def main() -> int:
    if not WORKBOOK.exists():
        print(f"Missing workbook: {WORKBOOK}", file=sys.stderr)
        return 2
    if not db_enabled():
        print("DATABASE_URL is not configured or psycopg is unavailable.", file=sys.stderr)
        return 2

    wb = load_workbook(WORKBOOK, data_only=True)
    details = wb["تفاصيل شهر 6"]
    ops = wb["العمليات المالية"]

    shipment_rows = [list(r) for r in details.iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, "", "الإجمالي")]
    ops_rows = [list(r) for r in ops.iter_rows(min_row=12, values_only=True) if r and r[0] not in (None, "")]

    shipments = [shipment_payload(row) for row in shipment_rows]
    wallet = [wallet_payload(row, "finance") for row in ops_rows]
    payments = [wallet_payload(row, "payment") for row in ops_rows]
    cod = [
        {
            "order_id": norm(row[0]),
            "tracking_number": norm(row[1]),
            "merchant_name": norm(row[2]),
            "customer_name": norm(row[4]),
            "cod_amount": num(row[7]),
            "collection_date": datev(row[13]),
            "transfer_date": None,
            "settlement_status": norm(row[14]),
            "shipment_status": norm(row[12]),
            "raw_payload": row,
        }
        for row in shipment_rows
    ]

    with get_conn() as conn:
        ensure_schema(conn)
        run_id = make_sync_run(conn, "import_report_to_pg.py")
        s_ins, s_upd = upsert_rows(conn, "shipments", shipments, ["order_id"], [
            "tracking_number","merchant_name","store_name","customer_name","city","carrier","payment_type","status",
            "shipment_date","delivery_date","weight","cod_amount","shipping_charge","customer_price_gross",
            "customer_price_net","platform_cost_gross","platform_cost_net","base_profit","extra_kg","extra_profit",
            "cod_profit","total_profit","included_in_profit","source_row","source_hash","raw_payload"
        ])
        w_ins, w_upd = upsert_rows(conn, "wallet_transactions", wallet, ["transaction_key"], [
            "transaction_date","user_name","description","amount","transaction_type","balance_before","balance_after","source_page","raw_payload"
        ])
        p_ins, p_upd = upsert_rows(conn, "payments", payments, ["payment_key"], [
            "payment_date","customer_name","amount","method","status","source_page","raw_payload"
        ])
        c_ins, c_upd = upsert_rows(conn, "cod_collections", cod, ["order_id"], [
            "tracking_number","merchant_name","customer_name","cod_amount","collection_date","transfer_date","settlement_status","shipment_status","raw_payload"
        ])
        finish_sync_run(conn, run_id, "ok", inserted=s_ins+w_ins+p_ins+c_ins, updated=s_upd+w_upd+p_upd+c_upd)

    print(json.dumps({
        "source_workbook": str(WORKBOOK),
        "shipments": len(shipments),
        "wallet": len(wallet),
        "payments": len(payments),
        "cod": len(cod),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
