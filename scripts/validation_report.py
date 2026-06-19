from __future__ import annotations

import json
import math
import random
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from scripts.db_store import (
    aggregate_shipments,
    aggregate_top_entities,
    db_enabled,
    get_conn,
    shipment_samples,
)


OUTPUT = PROJECT_DIR / "validation_report.json"
WORKBOOK = PROJECT_DIR / "output" / "lead6_report.xlsx"


def pct_diff(current: float, new: float) -> float:
    base = abs(current) if current not in (0, 0.0) else 1.0
    return abs(new - current) / base * 100.0


def money(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def workbook_metrics():
    wb = load_workbook(WORKBOOK, data_only=True)
    details = wb["تفاصيل شهر 6"]
    summary = wb["ملخصات"]
    rows = list(details.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r and r[0] not in (None, "", "الإجمالي")]
    metrics = {
        "shipments": len(rows),
        "revenue": money(summary["B6"].value if False else None),
    }
    metrics["revenue"] = sum(money(r[16]) for r in rows)
    metrics["platform_cost"] = sum(money(r[17]) for r in rows)
    metrics["overweight_profit"] = sum(money(r[20]) for r in rows)
    metrics["cod_profit"] = sum(money(r[21]) for r in rows)
    metrics["total_profit"] = sum(money(r[22]) for r in rows)
    metrics["top_merchants"] = {}
    metrics["top_cities"] = {}
    metrics["wallets"] = None
    metrics["payments"] = None
    return metrics, rows


def compare_metric(name, current, new, reason=""):
    current = float(current)
    new = float(new)
    diff = new - current
    return {
        "metric_name": name,
        "current_value": current,
        "new_value": new,
        "difference": diff,
        "difference_pct": pct_diff(current, new),
        "reason": reason,
        "ok": pct_diff(current, new) <= 0.01,
    }


def main() -> int:
    if not db_enabled():
        report = {
            "status": "pending",
            "reason": "DATABASE_URL not configured",
            "spec": "See validation_report_spec.md",
        }
        OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 0

    workbook, rows = workbook_metrics()
    with get_conn() as conn:
        pg = aggregate_shipments(conn)
        top_merchants = aggregate_top_entities(conn, "shipments", "merchant_name")
        top_cities = aggregate_top_entities(conn, "shipments", "city")
        sample_ids = [str(r[0]) for r in rows if r and r[0] not in (None, "")]
        random.seed(42)
        sample_ids = random.sample(sample_ids, k=min(20, len(sample_ids)))
        pg_samples = shipment_samples(conn, sample_ids)

    metrics = [
        compare_metric("عدد الشحنات", workbook["shipments"], pg["total_count"]),
        compare_metric("الإيرادات", workbook["revenue"], pg["cod_amount"]),
        compare_metric("تكلفة المنصة", workbook["platform_cost"], 0.0, "DB currently stores shipment-level platform fields only"),
        compare_metric("ربح الوزن الزائد", workbook["overweight_profit"], pg["extra_total"]),
        compare_metric("ربح COD", workbook["cod_profit"], pg["cod_total"]),
        compare_metric("إجمالي الربح", workbook["total_profit"], pg["total_profit"]),
    ]
    differences = [m for m in metrics if not m["ok"]]

    samples = []
    for order_id in sample_ids:
        old = next((r for r in rows if str(r[0]) == order_id), None)
        new = pg_samples.get(order_id)
        if not old or not new:
            samples.append({"order_id": order_id, "status": "missing", "old": bool(old), "new": bool(new)})
            continue
        old_payload = {
            "base_profit": money(old[19]),
            "extra_profit": money(old[20]),
            "cod_profit": money(old[21]),
            "total_profit": money(old[22]),
        }
        new_payload = {
            "base_profit": money(new.get("base_profit")),
            "extra_profit": money(new.get("extra_profit")),
            "cod_profit": money(new.get("cod_profit")),
            "total_profit": money(new.get("total_profit")),
        }
        samples.append({
            "order_id": order_id,
            "old": old_payload,
            "new": new_payload,
            "ok": all(abs(old_payload[k] - new_payload[k]) <= 0.01 for k in old_payload),
        })

    report = {
        "status": "ok" if not differences else "mismatch",
        "metrics": metrics,
        "top_merchants": top_merchants,
        "top_cities": top_cities,
        "sample_size": len(samples),
        "samples": samples,
        "differences": differences,
        "policy": "PostgreSQL remains non-authoritative until all metrics are within 0.01%.",
    }
    OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
