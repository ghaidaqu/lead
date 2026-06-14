from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path("/Users/ghaida/Desktop/lead")
SOURCE = ROOT / "output" / "lead6_report.xlsx"

PALETTE = {
    "dark": "5B4A47",
    "navy": "2F2628",
    "cream": "F7F1EA",
    "cream2": "FFFDFB",
    "sand": "EDE3D8",
    "rose": "F4E3E0",
    "rose2": "FFF1F0",
    "teal": "8B746D",
    "muted": "7C6B68",
    "line": "E1D4CA",
    "white": "FFFFFF",
}


def fill(color):
    return PatternFill("solid", fgColor=color)


def style_range(ws, min_row, max_row, min_col, max_col, *, fill_color=None, font=None, border=None, alignment=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def set_tab(ws, color):
    ws.sheet_properties.tabColor = color
    ws.sheet_view.showGridLines = False
    ws.sheet_view.rightToLeft = True


def main():
    wb = load_workbook(SOURCE)

    thin = Side(style="thin", color=PALETTE["line"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)

    title_font = Font(name="Segoe UI", color=PALETTE["white"], bold=True, size=15)
    section_font = Font(name="Segoe UI", color=PALETTE["white"], bold=True, size=12)
    header_font = Font(name="Segoe UI", color=PALETTE["white"], bold=True, size=11)
    label_font = Font(name="Segoe UI", color=PALETTE["dark"], bold=True, size=11)
    body_font = Font(name="Segoe UI", color=PALETTE["navy"], size=10)
    bold_body = Font(name="Segoe UI", color=PALETTE["navy"], bold=True, size=10)

    # Summary sheet
    ws = wb["ملخصات"]
    set_tab(ws, PALETTE["dark"])
    ws["A1"].fill = fill(PALETTE["dark"])
    ws["A1"].font = title_font
    ws["A1"].alignment = right
    style_range(ws, 3, 50, 1, 2, border=border, alignment=right)
    for r in range(3, ws.max_row + 1):
        label = ws.cell(r, 1).value
        value = ws.cell(r, 2).value
        if label is None and value is None:
            continue
        if r == 3:
            ws.cell(r, 1).fill = fill(PALETTE["navy"])
            ws.cell(r, 1).font = section_font
            ws.cell(r, 2).fill = fill(PALETTE["navy"])
            ws.cell(r, 2).font = section_font
            ws.cell(r, 1).alignment = right
            ws.cell(r, 2).alignment = center
            continue
        if r in (19, 33):
            for c in range(1, 7):
                ws.cell(r, c).fill = fill(PALETTE["navy"])
                ws.cell(r, c).font = header_font
                ws.cell(r, c).alignment = center if c > 1 else right
                ws.cell(r, c).border = border
            continue
        ws.cell(r, 1).fill = fill(PALETTE["cream"])
        ws.cell(r, 1).font = label_font
        ws.cell(r, 1).alignment = right
        ws.cell(r, 2).fill = fill(PALETTE["cream2"])
        ws.cell(r, 2).font = bold_body
        ws.cell(r, 2).alignment = center
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.sheet_view.zoomScale = 90

    # Detail sheet
    ws = wb["تفاصيل شهر 6"]
    set_tab(ws, PALETTE["teal"])
    for cell in ws[1]:
        cell.fill = fill(PALETTE["navy"])
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in range(2, ws.max_row + 1):
        fill_color = PALETTE["cream2"] if r % 2 == 0 else PALETTE["cream"]
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = fill(fill_color)
            cell.font = body_font
            cell.border = border
            cell.alignment = right if c in (1, 2, 3, 4, 5, 6) else center
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 85

    # Prices sheet
    ws = wb["الاسعار"]
    set_tab(ws, PALETTE["sand"])
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = fill(PALETTE["navy"])
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if c == 1:
                cell.fill = fill(PALETTE["cream"])
                cell.font = label_font
                cell.alignment = right
            else:
                if r in (2, 4):
                    cell.fill = fill(PALETTE["sand"])
                    cell.font = bold_body
                else:
                    cell.fill = fill(PALETTE["cream2"])
                    cell.font = body_font
                cell.alignment = center
    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 100

    # Financial operations
    ws = wb["العمليات المالية"]
    set_tab(ws, PALETTE["dark"])
    ws["A1"].fill = fill(PALETTE["dark"])
    ws["A1"].font = title_font
    ws["A1"].alignment = right
    for r in range(3, 7):
        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.border = border
            if r == 3:
                cell.fill = fill(PALETTE["navy"])
                cell.font = section_font
                cell.alignment = center
            else:
                cell.fill = fill(PALETTE["cream"] if c != 3 else PALETTE["cream2"])
                cell.font = bold_body if c == 3 else label_font
                cell.alignment = center if c != 1 else right
    for c in range(1, 9):
        cell = ws.cell(8, c)
        cell.fill = fill(PALETTE["navy"])
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    for r in range(9, ws.max_row + 1):
        channel = str(ws.cell(r, 8).value or "")
        if not any(ws.cell(r, c).value is not None for c in range(1, 9)):
            continue
        if channel == "البنك":
            row_fill = PALETTE["cream"]
        elif channel == "ميسر":
            row_fill = PALETTE["cream2"]
        else:
            row_fill = PALETTE["sand"]
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.fill = fill(row_fill)
            cell.font = body_font
            cell.border = border
            cell.alignment = right if c in (1, 2, 3, 4, 5, 6, 7, 8) else center
    ws.freeze_panes = "A8"
    ws.sheet_view.zoomScale = 90

    # Statement sheet
    ws = wb["كشف الحساب"]
    set_tab(ws, PALETTE["rose"])
    for rng in ["A1:F1", "A2:F2", "A3:B3", "D3:E3"]:
        top_left = ws[rng.split(":")[0]]
        top_left.fill = fill(PALETTE["dark"] if rng != "A2:F2" else PALETTE["cream"])
        top_left.border = border
    ws["A1"].font = title_font
    ws["A1"].alignment = right
    ws["A2"].font = Font(name="Segoe UI", color=PALETTE["muted"], italic=True, size=10)
    ws["A2"].alignment = right
    for cell_ref in ["A3", "D3"]:
        ws[cell_ref].fill = fill(PALETTE["navy"])
        ws[cell_ref].font = section_font
        ws[cell_ref].alignment = right
    for r in [4, 5, 6, 7]:
        ws.cell(r, 1).fill = fill(PALETTE["cream"])
        ws.cell(r, 1).font = label_font
        ws.cell(r, 1).border = border
        ws.cell(r, 1).alignment = right
        ws.cell(r, 2).fill = fill(PALETTE["cream2"])
        ws.cell(r, 2).font = bold_body
        ws.cell(r, 2).border = border
        ws.cell(r, 2).alignment = center
    for r in range(9, ws.max_row + 1):
        if not any(ws.cell(r, c).value is not None for c in range(1, 4)):
            continue
        fill_color = PALETTE["rose2"] if r % 2 == 0 else PALETTE["cream"]
        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.fill = fill(fill_color)
            cell.font = body_font
            cell.border = border
            cell.alignment = wrap if c == 2 else center
    ws.freeze_panes = "A9"
    ws.sheet_view.zoomScale = 95
    ws.sheet_view.rightToLeft = True

    # Clean up any lingering blue fill artifacts in summary sheets.
    for sheet_name in ["ملخصات", "العمليات المالية"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                fg = cell.fill.fgColor
                if fg and fg.type == "rgb" and fg.rgb and str(fg.rgb).upper() in {"00DDEBF7", "0016324F"}:
                    cell.fill = fill(PALETTE["navy"])
                    cell.font = header_font if cell.row in (3, 19, 33, 8) else body_font
                    cell.alignment = center if cell.col_idx > 1 else right

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_view.showGridLines:
            ws.sheet_view.showGridLines = False

    wb.save(SOURCE)


if __name__ == "__main__":
    main()
