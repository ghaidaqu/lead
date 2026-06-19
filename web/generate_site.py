import html
import json
import os
import re
import sys
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from scripts.db_store import db_enabled, fetch_dashboard_rows, get_conn


SOURCE = PROJECT_DIR / "output" / "lead6_report.xlsx"
OUT_DIR = Path(__file__).resolve().parent
INDEX = OUT_DIR / "index.html"


def clean(value):
    return str(value).strip() if value is not None else ""


def money(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = clean(value).replace(",", "")
    num = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
    if num in ("", "-", ".", "-."):
        return 0.0
    return float(num)


def june_period_label(dates):
    days = set()
    for date in dates:
        if not date:
            continue
        try:
            if isinstance(date, datetime):
                days.add(date.day)
            else:
                days.add(datetime.fromisoformat(str(date)).day)
        except Exception:
            continue
    days = sorted(days)
    if not days:
        return "يونيو"
    if days[-1] == 1:
        return "1 يونيو"
    return f"1-{days[-1]} يونيو"


SPECIAL_MERCHANTS = {"عبدالرحمن المطيري", "مؤسسة اواني القيصرية التجارية"}
CARRIER_PRICE_NAMES = {
    "ارامكس - ARAMEX": "ارامكس",
    "aramex ( استلام من الفرع )": "ارامكس استلام",
    "SMSA - سمسا": "سمسا",
    "SMSA ( استلام من الفرع )": "سمسا استلام",
    "RedBox - ريدبوكس": "ريد بوكس",
    "JT Express": "JT Express",
}
RETURN_DETAIL_OVERRIDES = {
    "767": {
        "merchant": "مؤسسة اواني القيصرية التجارية",
        "carrier": "SMSA - سمسا",
        "weight": 15.0,
    },
    "1376": {
        "merchant": "عبدالرحمن المطيري",
        "carrier": "ارامكس - ARAMEX",
        "weight": 1.0,
    },
    "1389": {
        "merchant": "ليد إكسبرس",
        "carrier": "ارامكس - ARAMEX",
        "weight": 1.0,
    },
    "1256": {
        "merchant": "ليد إكسبرس",
        "carrier": "ارامكس - ARAMEX",
        "weight": 22.0,
    },
    "1315": {
        "merchant": "عبدالرحمن المطيري",
        "carrier": "SMSA - سمسا",
        "weight": 1.0,
    },
    "1356": {
        "merchant": "ليد إكسبرس",
        "carrier": "ارامكس - ARAMEX",
        "weight": 10.0,
    },
}


def price_sheet_headers(ws):
    if ws is None:
        return {}
    headers = {clean(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    normalized = {}
    for key, col in headers.items():
        norm = re.sub(r"\s+", "", key).replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").lower()
        normalized[norm] = col
    return headers | {f"__normalized__{k}": v for k, v in normalized.items()}


def header_lookup(headers, *candidates):
    normalized = {}
    for key, col in headers.items():
        if key.startswith("__normalized__"):
            normalized[key.replace("__normalized__", "")] = col
    for candidate in candidates:
        cand = re.sub(r"\s+", "", candidate).replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").lower()
        if cand in normalized:
            return normalized[cand]
    for key, col in normalized.items():
        for candidate in candidates:
            cand = re.sub(r"\s+", "", candidate).replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").lower()
            if cand and (cand in key or key in cand):
                return col
    return None


def platform_cost_net_from_prices(prices_ws, headers, merchant, carrier, fallback=0.0):
    if prices_ws is None:
        return fallback
    if merchant in SPECIAL_MERCHANTS and merchant in headers:
        return money(prices_ws.cell(5, headers[merchant]).value) or fallback
    carrier_key = CARRIER_PRICE_NAMES.get(carrier)
    if carrier_key and carrier_key in headers:
        return money(prices_ws.cell(5, headers[carrier_key]).value) or fallback
    return fallback


def infer_return_profit_from_prices(prices_ws, headers, revenue):
    if prices_ws is None:
        revenue_net = revenue * 0.85
        return {
            "merchant": "مقدر",
            "carrier": "مقدر",
            "weight": 0.0,
            "revenue_net": revenue_net,
            "platform_shipping": 0.0,
            "base_profit": revenue_net,
            "extra_profit": 0.0,
            "total_profit": revenue_net,
            "status": "مقدر بدون شيت الأسعار",
        }

    extra_col = header_lookup(headers, "السعر لكل كيلو  زيادة", "السعر لكل كيلو زيادة", "سعر لكل كيلو زيادة", "لكل كيلو زيادة", "extra kilo", "overweight")
    extra_customer_gross = money(prices_ws.cell(2, extra_col).value) if extra_col else 0.0
    extra_customer_net = money(prices_ws.cell(3, extra_col).value) if extra_col else 0.0
    extra_platform_gross = money(prices_ws.cell(4, extra_col).value) if extra_col else 0.0
    extra_platform_net = money(prices_ws.cell(5, extra_col).value) if extra_col else 0.0
    if not extra_customer_net and extra_customer_gross:
        extra_customer_net = extra_customer_gross * 0.85
    if not extra_platform_net and extra_platform_gross:
        extra_platform_net = extra_platform_gross * 0.85

    exact_candidates = []
    fallback_candidates = []
    skip_names = {"السعر لكل كيلو  زيادة", "السعر لكل كيلو زيادة", "سعر توصيل cod", ""}
    for name, col in headers.items():
        if name in skip_names:
            continue
        customer_gross = money(prices_ws.cell(2, col).value)
        platform_net = money(prices_ws.cell(5, col).value)
        if not customer_gross or not platform_net or revenue + 0.001 < customer_gross:
            continue

        customer_net = money(prices_ws.cell(3, col).value)
        if not customer_net:
            customer_net = customer_gross if name in SPECIAL_MERCHANTS else customer_gross * 0.85

        extra_charge = max(revenue - customer_gross, 0.0)
        extra_kg = extra_charge / extra_customer_gross if extra_customer_gross else 0.0
        revenue_net = customer_net + (extra_kg * extra_customer_net)
        platform_shipping = platform_net + (extra_kg * extra_platform_net)
        base_profit = customer_net - platform_net
        extra_profit = extra_kg * (extra_customer_net - extra_platform_net)
        total_profit = revenue_net - platform_shipping
        row = {
            "merchant": name if name in SPECIAL_MERCHANTS else "مقدر",
            "carrier": name if name not in SPECIAL_MERCHANTS else "مقدر",
            "weight": 15 + extra_kg if extra_kg else 0.0,
            "revenue_net": revenue_net,
            "platform_shipping": platform_shipping,
            "base_profit": base_profit,
            "extra_profit": extra_profit,
            "total_profit": total_profit,
            "status": f"مقدر حسب {name}",
        }
        if abs(revenue - customer_gross) < 0.001:
            exact_candidates.append(row)
        else:
            fallback_candidates.append(row)

    candidates = exact_candidates or fallback_candidates
    if not candidates:
        revenue_net = revenue * 0.85
        return {
            "merchant": "مقدر",
            "carrier": "غير موجود في شيت الأسعار",
            "weight": 0.0,
            "revenue_net": revenue_net,
            "platform_shipping": 0.0,
            "base_profit": revenue_net,
            "extra_profit": 0.0,
            "total_profit": revenue_net,
            "status": "مقدر بدون مطابقة سعر",
        }
    return min(candidates, key=lambda item: item["total_profit"])


def return_profit_from_details(prices_ws, headers, revenue, merchant, carrier, weight):
    if prices_ws is None:
        return infer_return_profit_from_prices(prices_ws, headers, revenue)

    extra_col = header_lookup(headers, "السعر لكل كيلو  زيادة", "السعر لكل كيلو زيادة", "سعر لكل كيلو زيادة", "لكل كيلو زيادة", "extra kilo", "overweight")
    extra_customer_gross = money(prices_ws.cell(2, extra_col).value) if extra_col else 0.0
    extra_platform_gross = money(prices_ws.cell(4, extra_col).value) if extra_col else 0.0
    extra_platform_net = money(prices_ws.cell(5, extra_col).value) if extra_col else 0.0
    if not extra_platform_net and extra_platform_gross:
        extra_platform_net = extra_platform_gross * 0.85

    price_key = merchant if merchant in SPECIAL_MERCHANTS and merchant in headers else CARRIER_PRICE_NAMES.get(carrier, carrier)
    col = headers.get(price_key)
    if col is None:
        inferred = infer_return_profit_from_prices(prices_ws, headers, revenue)
        inferred.update({
            "merchant": merchant,
            "carrier": carrier,
            "weight": weight,
            "status": "محسوب بمطابقة سعر تقديرية",
        })
        return inferred

    base_customer_gross = money(prices_ws.cell(2, col).value)
    platform_base_net = money(prices_ws.cell(5, col).value)
    extra_kg = max(weight - 15, 0.0)
    extra_charge_gross = max(revenue - base_customer_gross, 0.0)
    extra_charge_net = extra_charge_gross if merchant in SPECIAL_MERCHANTS else extra_charge_gross * 0.85
    revenue_net = revenue if merchant in SPECIAL_MERCHANTS else revenue * 0.85
    platform_extra_net = extra_kg * extra_platform_net
    platform_shipping = platform_base_net + platform_extra_net
    total_profit = revenue_net - platform_shipping
    extra_profit = extra_charge_net - platform_extra_net
    base_profit = total_profit - extra_profit
    return {
        "merchant": merchant,
        "carrier": carrier,
        "weight": weight,
        "revenue_net": revenue_net,
        "platform_shipping": platform_shipping,
        "base_profit": base_profit,
        "extra_profit": extra_profit,
        "total_profit": total_profit,
        "status": "محسوب من المحفظة والشحنات",
    }


def load_data():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    detail_sheet = None
    for candidate in ("تفاصيل شهر 6", "تفاصيل شهر 06", "تفاصيل يونيو", "تفاصيل", "الشحنات", "Raw_Shipments"):
        if candidate in wb.sheetnames:
            detail_sheet = candidate
            break
    if detail_sheet is None:
        detail_sheet = wb.sheetnames[0]
    ws = wb[detail_sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    ops = wb["العمليات المالية"] if "العمليات المالية" in wb.sheetnames else None
    stmt = wb["كشف الحساب"] if "كشف الحساب" in wb.sheetnames else None
    prices_ws = wb["الاسعار"] if "الاسعار" in wb.sheetnames else None
    prices_headers = price_sheet_headers(prices_ws)

    items = []
    cod_items = []
    by_merchant = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_city = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_carrier = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_status = defaultdict(int)
    by_date = defaultdict(float)
    by_date_revenue = defaultdict(float)
    by_date_count = defaultdict(int)
    all_shipment_dates = set()
    finance = {
        "bank": {"count": 0, "total": 0.0},
        "moyasar": {"count": 0, "total": 0.0},
        "shipping_return": {"count": 0, "total": 0.0},
        "tax_deduction": {"count": 0, "total": 0.0},
        "shipping_refund": {"count": 0, "total": 0.0},
        "other_net": 0.0,
        "other": {"count": 0, "total": 0.0},
    }
    statement = {
        "summary": {},
        "rows": [],
    }
    total_customer_shipping = 0.0
    total_records = 0
    all_items_by_order = {}
    return_items = []

    for row in rows:
        if not any(v is not None for v in row):
            continue
        if row[0] in (None, ""):
            continue
        total_records += 1
        item = {
            "order_id": row[0],
            "tracking": clean(row[1]),
            "merchant": clean(row[2]),
            "customer": clean(row[4]),
            "city": clean(row[5]),
            "carrier": clean(row[10]),
            "date": row[13].date().isoformat() if isinstance(row[13], datetime) else clean(row[13]),
            "status": clean(row[12]),
            "weight": money(row[11]),
            "included": clean(row[14]) == "نعم",
            "customer_shipping": money(row[6]),
            "cod_amount": money(row[7]),
            "customer_gross": money(row[15]),
            "customer_net": money(row[16]),
            "platform_shipping": money(row[17]),
            "shipping_profit": money(row[19]),
            "extra_profit": money(row[20]),
            "fee_profit": money(row[21]),
            "total_profit": money(row[22]),
            "review_diff": money(row[19]),
        }
        if item["order_id"] not in (None, ""):
            total_customer_shipping += item["customer_shipping"]
            all_items_by_order[str(item["order_id"])] = item
        if item["date"]:
            all_shipment_dates.add(item["date"])
        if not item["included"]:
            continue
        items.append(item)
        if item["cod_amount"] > 0:
            cod_items.append(item)
        by_merchant[item["merchant"]]["count"] += 1
        by_merchant[item["merchant"]]["total"] += item["total_profit"]
        by_city[item["city"]]["count"] += 1
        by_city[item["city"]]["total"] += item["total_profit"]
        by_carrier[item["carrier"]]["count"] += 1
        by_carrier[item["carrier"]]["total"] += item["total_profit"]
        by_status[item["status"]] += 1
        if item["date"]:
            by_date[item["date"]] += item["total_profit"]
            by_date_revenue[item["date"]] += item["customer_shipping"]
            by_date_count[item["date"]] += 1

    if ops is not None:
        for r in range(2, ops.max_row + 1):
            typ = clean(ops.cell(r, 4).value)
            note = clean(ops.cell(r, 6).value)
            amount = money(ops.cell(r, 5).value)
            if typ == "إيداع":
                key = "bank" if "تحويل بنكي" in note else "moyasar" if "Moyasar" in note else "other"
                finance[key]["count"] += 1
                finance[key]["total"] += amount
            else:
                if typ == "admin_deduction" and "خصم تكلفة شحنة مرتجعة" in note:
                    finance["shipping_return"]["count"] += 1
                    finance["shipping_return"]["total"] += abs(amount)
                    order_match = re.search(r"طلب\s*#?\s*(\d+)", note)
                    order_id = order_match.group(1) if order_match else ""
                    linked_item = all_items_by_order.get(order_id)
                    revenue = abs(amount)
                    if linked_item is not None:
                        net_ratio = (
                            linked_item["customer_net"] / linked_item["customer_gross"]
                            if linked_item["customer_gross"]
                            else 0.85
                        )
                        revenue_net = revenue * net_ratio
                        platform_cost_net = platform_cost_net_from_prices(
                            prices_ws,
                            prices_headers,
                            linked_item["merchant"],
                            linked_item["carrier"],
                            linked_item["platform_shipping"],
                        )
                        base_profit = revenue_net - platform_cost_net
                        extra_profit = linked_item["extra_profit"]
                        total_profit = base_profit + extra_profit
                        return_items.append({
                            "order_id": order_id,
                            "merchant": linked_item["merchant"],
                            "carrier": linked_item["carrier"],
                            "weight": linked_item["weight"],
                            "revenue": revenue,
                            "revenue_net": revenue_net,
                            "platform_shipping": platform_cost_net,
                            "base_profit": base_profit,
                            "extra_profit": extra_profit,
                            "total_profit": total_profit,
                            "matched": True,
                            "status": "محسوب",
                        })
                    else:
                        override = RETURN_DETAIL_OVERRIDES.get(order_id)
                        if override:
                            inferred = return_profit_from_details(
                                prices_ws,
                                prices_headers,
                                revenue,
                                override["merchant"],
                                override["carrier"],
                                override["weight"],
                            )
                        else:
                            inferred = infer_return_profit_from_prices(prices_ws, prices_headers, revenue)
                        return_items.append({
                            "order_id": order_id or "-",
                            "merchant": inferred["merchant"],
                            "carrier": inferred["carrier"],
                            "weight": inferred["weight"],
                            "revenue": revenue,
                            "revenue_net": inferred["revenue_net"],
                            "platform_shipping": inferred["platform_shipping"],
                            "base_profit": inferred["base_profit"],
                            "extra_profit": inferred["extra_profit"],
                            "total_profit": inferred["total_profit"],
                            "matched": False,
                            "status": inferred["status"],
                        })
                elif typ == "admin_deduction" and "خصم ضريبة القيمة المضافة" in note:
                    finance["other_net"] += amount
                    finance["tax_deduction"]["count"] += 1
                    finance["tax_deduction"]["total"] += abs(amount)
                elif typ == "استرداد" and "استرداد تكلفة شحن" in note:
                    finance["other_net"] += amount
                    finance["shipping_refund"]["count"] += 1
                    finance["shipping_refund"]["total"] += amount
                else:
                    finance["other_net"] += amount
                    finance["other"]["count"] += 1
                    finance["other"]["total"] += amount
    finance["total"] = finance["bank"]["total"] + finance["moyasar"]["total"]

    if stmt is not None:
        statement["summary"]["deposits_total"] = money(stmt["B4"].value)
        statement["summary"]["expenses_total"] = money(stmt["B5"].value)
        statement["summary"]["net_total"] = money(stmt["B6"].value)
        statement["summary"]["transfer_fees_total"] = money(stmt["B7"].value)
        statement["summary"]["deposits_count"] = money(stmt["E4"].value)
        statement["summary"]["expenses_count"] = money(stmt["E5"].value)
        statement["summary"]["period"] = clean(stmt["E6"].value)
        for r in range(10, stmt.max_row + 1):
            date_val = stmt.cell(r, 1).value
            expense_type = clean(stmt.cell(r, 2).value)
            amount = money(stmt.cell(r, 3).value)
            if not any([date_val, expense_type, amount]):
                continue
            statement["rows"].append({
                "date": date_val.date().isoformat() if isinstance(date_val, datetime) else clean(date_val),
                "expense_type": expense_type,
                "amount": amount,
            })

    items.sort(key=lambda x: (x["date"], x["order_id"]), reverse=True)
    top_merchants = sorted(by_merchant.items(), key=lambda kv: kv[1]["total"], reverse=True)[:5]
    top_cities = sorted(by_city.items(), key=lambda kv: kv[1]["total"], reverse=True)[:5]
    top_carriers = sorted(by_carrier.items(), key=lambda kv: kv[1]["total"], reverse=True)
    daily = sorted(by_date.items(), key=lambda kv: kv[0])
    daily_revenue = sorted(by_date_revenue.items(), key=lambda kv: kv[0])
    daily_count = sorted(by_date_count.items(), key=lambda kv: kv[0])
    period_label = june_period_label(all_shipment_dates)
    return_revenue = sum(item["revenue"] for item in return_items)
    return_profit = sum(item["total_profit"] for item in return_items)
    totals = {
        "records": total_records,
        "active": len(items),
        "shipping": len(items),
        "cod": sum(1 for item in items if item["fee_profit"] > 0),
        "cod_amount": sum(item["cod_amount"] for item in items if item["included"]),
        "revenue": total_customer_shipping + return_revenue,
        "base": sum(item["shipping_profit"] for item in items),
        "extra": sum(item["extra_profit"] for item in items),
        "cod_profit": sum(item["fee_profit"] for item in items),
        "return_revenue": return_revenue,
        "return_profit": return_profit,
        "return_count": len(return_items),
        "total": sum(item["total_profit"] for item in items) + return_profit,
        "excluded": total_records - len(items),
    }
    cod_items.sort(key=lambda x: (x["date"], x["order_id"]), reverse=True)
    return_items.sort(key=lambda x: (x["matched"], x["order_id"]), reverse=True)
    order_numbers = sorted(
        {
            int(item["order_id"])
            for item in all_items_by_order.values()
            if str(item["order_id"]).isdigit()
        }
    )
    missing_sequence_numbers = []
    if order_numbers:
        expected = set(range(order_numbers[0], order_numbers[-1] + 1))
        missing_sequence_numbers = sorted(expected.difference(order_numbers))
    return totals, top_merchants, top_cities, top_carriers, by_status, daily, daily_revenue, daily_count, finance, cod_items, return_items, statement, period_label, missing_sequence_numbers


def _row_value(row, index, default=None):
    try:
        value = row[index]
    except Exception:
        return default
    return default if value is None else value


def _row_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return clean(value)


def load_data_from_db():
    with get_conn() as conn:
        payload = fetch_dashboard_rows(conn)

    shipment_rows = payload["shipments"]
    wallet_rows = payload["wallet"]
    cod_rows_db = payload["cod"]

    items = []
    cod_items = []
    by_merchant = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_city = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_carrier = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_status = defaultdict(int)
    by_date = defaultdict(float)
    by_date_revenue = defaultdict(float)
    by_date_count = defaultdict(int)
    all_shipment_dates = set()
    all_items_by_order = {}
    finance = {
        "bank": {"count": 0, "total": 0.0},
        "moyasar": {"count": 0, "total": 0.0},
        "shipping_return": {"count": 0, "total": 0.0},
        "tax_deduction": {"count": 0, "total": 0.0},
        "shipping_refund": {"count": 0, "total": 0.0},
        "other_net": 0.0,
        "other": {"count": 0, "total": 0.0},
    }
    statement = {"summary": {}, "rows": []}
    return_items = []
    total_customer_shipping = 0.0

    for row in shipment_rows:
        raw = row.get("raw_payload") or []
        if not raw:
            raw = [
                row.get("order_id"),
                row.get("tracking_number"),
                row.get("merchant_name"),
                row.get("store_name"),
                row.get("customer_name"),
                row.get("city"),
                row.get("shipping_charge"),
                row.get("cod_amount"),
                None,
                row.get("payment_type"),
                row.get("carrier"),
                row.get("weight"),
                row.get("status"),
                row.get("shipment_date"),
                "نعم" if row.get("included_in_profit") else "لا",
                row.get("customer_price_gross"),
                row.get("customer_price_net"),
                row.get("platform_cost_net"),
                row.get("extra_kg"),
                row.get("base_profit"),
                row.get("extra_profit"),
                row.get("cod_profit"),
                row.get("total_profit"),
                row.get("source_row"),
            ]
        if not any(v is not None for v in raw):
            continue
        if raw[0] in (None, ""):
            continue
        item = {
            "order_id": raw[0],
            "tracking": clean(_row_value(raw, 1)),
            "merchant": clean(_row_value(raw, 2)),
            "customer": clean(_row_value(raw, 4)),
            "city": clean(_row_value(raw, 5)),
            "carrier": clean(_row_value(raw, 10)),
            "date": _row_date(_row_value(raw, 13)),
            "status": clean(_row_value(raw, 12)),
            "weight": money(_row_value(raw, 11)),
            "included": clean(_row_value(raw, 14)) == "نعم",
            "customer_shipping": money(_row_value(raw, 6)),
            "cod_amount": money(_row_value(raw, 7)),
            "customer_gross": money(_row_value(raw, 15)),
            "customer_net": money(_row_value(raw, 16)),
            "platform_shipping": money(_row_value(raw, 17)),
            "shipping_profit": money(_row_value(raw, 19)),
            "extra_profit": money(_row_value(raw, 20)),
            "fee_profit": money(_row_value(raw, 21)),
            "total_profit": money(_row_value(raw, 22)),
            "review_diff": money(_row_value(raw, 19)),
        }
        total_customer_shipping += item["customer_shipping"]
        all_items_by_order[str(item["order_id"])] = item
        if item["date"]:
            all_shipment_dates.add(item["date"])
        if not item["included"]:
            continue
        items.append(item)
        if item["cod_amount"] > 0:
            cod_items.append(item)
        by_merchant[item["merchant"]]["count"] += 1
        by_merchant[item["merchant"]]["total"] += item["total_profit"]
        by_city[item["city"]]["count"] += 1
        by_city[item["city"]]["total"] += item["total_profit"]
        by_carrier[item["carrier"]]["count"] += 1
        by_carrier[item["carrier"]]["total"] += item["total_profit"]
        by_status[item["status"]] += 1
        if item["date"]:
            by_date[item["date"]] += item["total_profit"]
            by_date_revenue[item["date"]] += item["customer_shipping"]
            by_date_count[item["date"]] += 1

    for row in wallet_rows:
        raw = row.get("raw_payload") or []
        typ = clean(_row_value(raw, 3))
        note = clean(_row_value(raw, 5))
        amount = money(_row_value(raw, 4))
        if typ == "إيداع":
            key = "bank" if "تحويل بنكي" in note else "moyasar" if "Moyasar" in note else "other"
            finance[key]["count"] += 1
            finance[key]["total"] += amount
        else:
            if typ == "admin_deduction" and "خصم تكلفة شحنة مرتجعة" in note:
                finance["shipping_return"]["count"] += 1
                finance["shipping_return"]["total"] += abs(amount)
                order_match = re.search(r"طلب\s*#?\s*(\d+)", note)
                order_id = order_match.group(1) if order_match else ""
                linked_item = all_items_by_order.get(order_id)
                revenue = abs(amount)
                if linked_item is not None:
                    net_ratio = (linked_item["customer_net"] / linked_item["customer_gross"]) if linked_item["customer_gross"] else 0.85
                    revenue_net = revenue * net_ratio
                    platform_cost_net = linked_item["platform_shipping"]
                    base_profit = revenue_net - platform_cost_net
                    extra_profit = linked_item["extra_profit"]
                    total_profit = base_profit + extra_profit
                else:
                    revenue_net = revenue * 0.85
                    platform_cost_net = 0.0
                    base_profit = revenue_net
                    extra_profit = 0.0
                    total_profit = revenue_net
                return_items.append({
                    "order_id": order_id or "-",
                    "merchant": linked_item["merchant"] if linked_item else "مقدر",
                    "carrier": linked_item["carrier"] if linked_item else "مقدر",
                    "weight": linked_item["weight"] if linked_item else 0.0,
                    "revenue": revenue,
                    "revenue_net": revenue_net,
                    "platform_shipping": platform_cost_net,
                    "base_profit": base_profit,
                    "extra_profit": extra_profit,
                    "total_profit": total_profit,
                    "matched": linked_item is not None,
                    "status": "محسوب" if linked_item else "مقدر",
                })
            elif typ == "admin_deduction" and "خصم ضريبة القيمة المضافة" in note:
                finance["other_net"] += amount
                finance["tax_deduction"]["count"] += 1
                finance["tax_deduction"]["total"] += abs(amount)
            elif typ == "استرداد" and "استرداد تكلفة شحن" in note:
                finance["other_net"] += amount
                finance["shipping_refund"]["count"] += 1
                finance["shipping_refund"]["total"] += amount
            else:
                finance["other_net"] += amount
                finance["other"]["count"] += 1
                finance["other"]["total"] += amount

    finance["total"] = finance["bank"]["total"] + finance["moyasar"]["total"]
    top_merchants = sorted(by_merchant.items(), key=lambda kv: kv[1]["total"], reverse=True)[:5]
    top_cities = sorted(by_city.items(), key=lambda kv: kv[1]["total"], reverse=True)[:5]
    top_carriers = sorted(by_carrier.items(), key=lambda kv: kv[1]["total"], reverse=True)
    daily = sorted(by_date.items(), key=lambda kv: kv[0])
    daily_revenue = sorted(by_date_revenue.items(), key=lambda kv: kv[0])
    daily_count = sorted(by_date_count.items(), key=lambda kv: kv[0])
    period_label = june_period_label(all_shipment_dates)
    return_revenue = sum(item["revenue"] for item in return_items)
    return_profit = sum(item["total_profit"] for item in return_items)
    totals = {
        "records": len(shipment_rows),
        "active": len(items),
        "shipping": len(items),
        "cod": sum(1 for item in items if item["fee_profit"] > 0),
        "cod_amount": sum(item["cod_amount"] for item in items if item["included"]),
        "revenue": total_customer_shipping + return_revenue,
        "base": sum(item["shipping_profit"] for item in items),
        "extra": sum(item["extra_profit"] for item in items),
        "cod_profit": sum(item["fee_profit"] for item in items),
        "return_revenue": return_revenue,
        "return_profit": return_profit,
        "return_count": len(return_items),
        "total": sum(item["total_profit"] for item in items) + return_profit,
        "excluded": len(shipment_rows) - len(items),
    }
    cod_items.sort(key=lambda x: (x["date"], x["order_id"]), reverse=True)
    return_items.sort(key=lambda x: (x["matched"], x["order_id"]), reverse=True)
    order_numbers = sorted(
        {
            int(item["order_id"])
            for item in all_items_by_order.values()
            if str(item["order_id"]).isdigit()
        }
    )
    missing_sequence_numbers = []
    if order_numbers:
        expected = set(range(order_numbers[0], order_numbers[-1] + 1))
        missing_sequence_numbers = sorted(expected.difference(order_numbers))
    return totals, top_merchants, top_cities, top_carriers, by_status, daily, daily_revenue, daily_count, finance, cod_items, return_items, statement, period_label, missing_sequence_numbers


def fmt_money(value):
    return f"{value:,.2f} ريال"


def metric_card(label, value, note="", tone="navy"):
    return f"""
      <article class="metric-card tone-{tone}">
        <span class="metric-label">{label}</span>
        <strong>{value}</strong>
        <div class="metric-footer">{note}</div>
      </article>
    """


def build_html(data):
    totals = data["totals"]
    daily = data["daily"]
    daily_revenue = data["daily_revenue"]
    daily_count = data["daily_count"]
    finance = data["finance"]
    cod_items = data["cod_items"]
    return_items = data["return_items"]
    top_merchants = data["top_merchants"]
    top_cities = data["top_cities"]
    top_carriers = data["top_carriers"]
    period_label = data["period_label"]
    missing_sequence_numbers = data.get("missing_sequence_numbers", [])
    statement = data.get("statement", {"summary": {}, "rows": []})
    details_href = "/report.xlsx"
    last_updated = datetime.fromtimestamp(SOURCE.stat().st_mtime).strftime("%d %b %Y - %I:%M %p") if SOURCE.exists() else datetime.now().strftime("%d %b %Y - %I:%M %p")
    cod_overrides = {
        "1757": {"collection_date": "2026-06-12", "transfer_date": "2026-06-13"},
        "1756": {"collection_date": "2026-06-11", "transfer_date": "2026-06-13"},
    }

    daily_profit_points = daily[-7:]
    daily_revenue_map = {date: revenue for date, revenue in daily_revenue}
    daily_count_map = {date: count for date, count in daily_count}
    daily_chart_labels = [
        str(date)[5:] if len(str(date)) >= 10 else str(date)
        for date, _ in daily_profit_points
    ]
    daily_chart_profit = [round(float(value), 2) for _, value in daily_profit_points]
    daily_chart_revenue = [
        round(float(daily_revenue_map.get(date, 0)), 2)
        for date, _ in daily_profit_points
    ]
    daily_chart_count = [
        int(daily_count_map.get(date, 0))
        for date, _ in daily_profit_points
    ]
    carrier_total = sum(data["total"] for _, data in top_carriers) or 1
    carrier_colors = ["#7A2438", "#B04A64", "#D8CABE", "#8FBF9F", "#C86B6B", "#6F6A66"]
    carrier_legend_items = []
    carrier_chart_labels = []
    carrier_chart_values = []
    carrier_chart_colors = []
    for idx, (name, data) in enumerate(top_carriers):
        pct = data["total"] / carrier_total
        color = carrier_colors[idx % len(carrier_colors)]
        display_name = "ARAMEX - ارامكس" if name == "ارامكس - ARAMEX" else name
        carrier_chart_labels.append(display_name)
        carrier_chart_values.append(round(float(data["total"]), 2))
        carrier_chart_colors.append(color)
        carrier_legend_items.append(
            f"""
              <div class="carrier-row">
                <span class="carrier-dot" style="background:{color}"></span>
                <div class="carrier-name">{html.escape(display_name)}</div>
                <div class="carrier-meta">{pct*100:.1f}%</div>
              </div>
            """
        )
    carrier_legend_html = "".join(carrier_legend_items)

    merchant_rows = "".join(
        f"<tr><td>{idx + 1}</td><td>{html.escape(name or '-')}</td><td>{data['count']}</td><td>{fmt_money(data['total'])}</td></tr>"
        for idx, (name, data) in enumerate(top_merchants[:4])
    )
    city_rows = "".join(
        f"<tr><td>{idx + 1}</td><td>{html.escape(name or '-')}</td><td>{data['count']}</td><td>{fmt_money(data['total'])}</td></tr>"
        for idx, (name, data) in enumerate(top_cities[:4])
    )
    profit_details = f"""
      <section class="profit-details-card">
        <article class="panel details-panel">
          <div class="panel-header compact-header">
            <div>
              <p class="eyebrow">تفاصيل الأرباح</p>
              <p class="subtle">تفصيل القيم الداخلة في إجمالي الربح</p>
            </div>
          </div>
          <div class="details-grid">
            <div class="detail-chip"><span>ربح الشحنات</span><strong>{fmt_money(totals["base"])}</strong></div>
            <div class="detail-chip"><span>ربح المرتجعات</span><strong>{fmt_money(totals["return_profit"])}</strong></div>
            <div class="detail-chip"><span>ربح الوزن الزائد</span><strong>{fmt_money(totals["extra"])}</strong></div>
            <div class="detail-chip"><span>ربح COD</span><strong>{fmt_money(totals["cod_profit"])}</strong></div>
            <div class="detail-chip"><span>عدد COD</span><strong>{totals["cod"]}</strong></div>
          </div>
        </article>
      </section>
    """
    finance_cards = f"""
      <section class="mini-band finance-band">
        <article class="mini-card finance-card bank">
          <span class="metric-label">إيداعات البنك</span>
          <strong>{finance['bank']['count']}</strong>
          <div class="metric-footer">{fmt_money(finance['bank']['total'])}</div>
        </article>
        <article class="mini-card finance-card moyasar">
          <span class="metric-label">إيداعات ميسر</span>
          <strong>{finance['moyasar']['count']}</strong>
          <div class="metric-footer">{fmt_money(finance['moyasar']['total'])}</div>
        </article>
        <article class="mini-card finance-card other">
          <span class="metric-label">تفصيل الخصومات والاستردادات</span>
          <div class="breakdown-line">خصم الضريبة: {fmt_money(finance['tax_deduction']['total'])}</div>
          <div class="breakdown-line">العمليات الملغية: {fmt_money(finance['shipping_refund']['total'])}</div>
        </article>
        <article class="mini-card finance-card finance-total-card other">
          <span class="metric-label">إجمالي الإيداعات</span>
          <strong>{fmt_money(finance['total'])}</strong>
          <div class="metric-footer">البنك + ميسر</div>
        </article>
      </section>
    """
    statement_rows = "".join(
        f"<tr class='expense-row'><td>{row['date']}</td><td>{row['expense_type'] or '-'}</td><td>{fmt_money(row['amount'])}</td></tr>"
        for row in statement["rows"]
    )

    def cod_collection_date(item):
        override = cod_overrides.get(str(item["order_id"])) or {}
        if override.get("collection_date"):
            return override["collection_date"]
        if item.get("date") == "2026-06-15":
            return "-"
        return item.get("date") or "-"

    cod_rows = "".join(
        f"<tr><td>{item['order_id']}</td><td>{html.escape(item['merchant'])}</td><td>{fmt_money(item['cod_amount'])}</td><td>{item.get('date') or '-'}</td><td>{cod_collection_date(item)}</td><td>{(cod_overrides.get(str(item['order_id'])) or {}).get('transfer_date') or '-'}</td></tr>"
        for item in cod_items
    )
    missing_sequence_text = "، ".join(str(num) for num in missing_sequence_numbers) if missing_sequence_numbers else "لا يوجد"
    return_rows = "".join(
        f"<tr><td>{item['order_id']}</td><td>{html.escape(item['merchant'])}</td><td>{html.escape(item['carrier'])}</td><td>{item['weight']:.2f} كجم</td><td>{fmt_money(item['revenue'])}</td><td>{fmt_money(item['platform_shipping'])}</td><td>{fmt_money(item['base_profit'])}</td><td>{fmt_money(item['extra_profit'])}</td><td>{fmt_money(item['total_profit'])}</td><td>{item['status']}</td></tr>"
        for item in return_items
    )
    return f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>lead</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
  <style>
    :root {{
      color-scheme: light;
      --bg: #F6F1EA;
      --sidebar: #FBF8F3;
      --card: #FFFFFF;
      --card-strong: #FFFFFF;
      --border: #E7DDD3;
      --text: #1C1C1C;
      --muted: #6F6A66;
      --accent: #7A2438;
      --accent-light: #B04A64;
      --beige: #D8CABE;
      --grid: #E8DED5;
      --positive: #587F63;
      --negative: #A24D4D;
      --shadow: 0 18px 45px rgba(45, 31, 25, .08);
    }}
    body.dark {{
      color-scheme: dark;
      --bg: #0F1113;
      --sidebar: #0B0D0F;
      --card: #171A1F;
      --card-strong: #1C2027;
      --border: #252932;
      --text: #F5F2EE;
      --muted: #A7A7A7;
      --accent: #7A2438;
      --accent-light: #B04A64;
      --beige: #D8CABE;
      --grid: #2A2D33;
      --positive: #8FBF9F;
      --negative: #C86B6B;
      --shadow: 0 18px 52px rgba(0,0,0,.28);
    }}
    *, *::before, *::after {{
      box-sizing: border-box;
      transition: background-color .35s ease, color .35s ease, border-color .35s ease, box-shadow .35s ease, fill .35s ease, stroke .35s ease;
    }}
    html {{ scroll-behavior:smooth; }}
    body {{
      margin:0;
      min-height:100vh;
      color:var(--text);
      font-family: Inter, "Avenir Next", "Segoe UI", "Noto Sans Arabic", system-ui, sans-serif;
      background:var(--bg);
    }}
    a {{ color:inherit; text-decoration:none; }}
    .page {{ width:min(1180px, calc(100% - 32px)); margin:0 auto; padding:20px 0 34px; }}
    .hero {{
      margin:0 0 12px;
      padding:12px 18px;
      border-radius:12px;
      background:linear-gradient(180deg, rgba(24,36,55,.95), rgba(14,24,40,.92));
      border:1px solid var(--border);
      box-shadow:0 16px 48px rgba(0,0,0,.34), inset 0 1px 0 rgba(255,255,255,.055);
      backdrop-filter: blur(6px) saturate(104%);
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:16px;
      flex-wrap:wrap;
      direction:rtl;
      position:relative;
      overflow:hidden;
    }}
    .hero::before, .hero::after {{ content:none; }}
    .brand-lockup {{ position:relative; z-index:1; display:flex; align-items:center; justify-content:flex-start; gap:0; padding:0; border:0; background:transparent; box-shadow:none; direction:ltr; flex:0 0 auto; }}
    .brand-logo {{ width:min(96px, 12vw); max-width:96px; height:auto; display:block; background:transparent; filter:none; }}
    .brand-divider {{ width:1px; height:38px; background:rgba(142,157,178,.30); box-shadow:0 0 18px rgba(122,36,56,.10); margin:0 8px 0 -2px; }}
    .hero-copy {{ display:grid; gap:0; justify-items:flex-start; margin-inline-start:-1px; }}
    .hero-copy span {{ display:block; margin:0; color:rgba(255,255,255,.92); font-size:clamp(14px, 1.6vw, 18px); line-height:1.0; font-weight:650; letter-spacing:-0.02em; }}
    .hero-copy .brand-tag {{ font-size:clamp(14px, 1.6vw, 18px); }}
    .hero-title {{
      position:relative;
      z-index:1;
      direction:rtl;
      text-align:right;
      display:flex;
      flex-direction:column;
      align-items:flex-end;
      justify-content:center;
      gap:3px;
      margin:0;
      padding:0;
      padding-left:0;
      padding-inline-start:0;
      padding-inline-end:0;
      min-width:0;
      width:auto;
      flex:0 1 auto;
    }}
    .hero-title h1 {{
      margin:0;
      font-size:clamp(17px, 1.8vw, 22px);
      line-height:1.08;
      letter-spacing:-0.03em;
      font-weight:800;
      color:#fff;
      text-shadow:0 2px 16px rgba(0,0,0,.32);
    }}
    .hero-title p {{
      margin:0;
      color:rgba(196,207,222,.72);
      font-size:11px;
    }}
    .hero-tools {{
      display:flex;
      align-items:center;
      justify-content:flex-end;
      gap:10px;
      flex-wrap:wrap;
      margin:0 0 10px;
    }}
    .badge {{
      display:inline-flex;
      align-items:center;
      min-height:40px;
      padding:0 14px;
      border-radius:10px;
      border:1px solid #E7DDD3;
      background:rgba(17,28,45,.78);
      color:rgba(230,238,248,.88);
      font-size:12px;
      backdrop-filter: blur(4px);
    }}
    .primary-button {{
      display:inline-flex;
      align-items:center;
      min-height:38px;
      padding:0 14px;
      border-radius:10px;
      background:
        linear-gradient(135deg, rgba(122,36,56,.20), rgba(176,74,100,.28)),
        linear-gradient(180deg, rgba(255,255,255,.12), rgba(255,255,255,.02));
      color:#fff;
      border:1px solid rgba(122,36,56,.22);
      font-weight:780;
      letter-spacing:-0.01em;
      box-shadow:
        0 14px 28px rgba(122,36,56,.08),
        inset 0 1px 0 rgba(255,255,255,.12);
      backdrop-filter: blur(4px);
      -webkit-backdrop-filter: blur(4px);
    }}
    .primary-button:hover {{ transform: translateY(-1px); filter: brightness(1.05); }}
    .subhero {{
      display:flex;
      justify-content:flex-end;
      margin:0 0 14px;
    }}
    .metric-grid {{ display:grid; grid-template-columns:repeat(4, minmax(0,1fr)); gap:10px; margin-bottom:10px; }}
    .mini-band {{ display:grid; grid-template-columns:repeat(3, minmax(0,1fr)); gap:14px; margin:0 0 16px; }}
    .finance-band {{ grid-template-columns:repeat(4, minmax(0,1fr)); }}
    .metric-card, .panel, .mini-card {{
      position:relative;
      overflow:hidden;
      border-radius:10px;
      background:linear-gradient(180deg, rgba(33,47,70,.95), rgba(17,29,46,.92));
      border:1px solid var(--border);
      box-shadow:0 16px 38px rgba(0,0,0,.32), inset 0 1px 0 rgba(255,255,255,.06);
      backdrop-filter: blur(6px) saturate(104%);
      -webkit-backdrop-filter: blur(6px) saturate(104%);
    }}
    .metric-card::before, .panel::before, .mini-card::before {{
      content:'';
      position:absolute;
      inset:0;
      background:linear-gradient(135deg, rgba(255,255,255,.045) 0%, rgba(255,255,255,.012) 30%, rgba(255,255,255,0) 58%);
      pointer-events:none;
    }}
    .metric-card > *, .panel > *, .mini-card > * {{ position:relative; z-index:1; }}
    .metric-card::after, .mini-card::after, .detail-chip::after {{
      content:'';
      position:absolute;
      top:0;
      left:14px;
      right:14px;
      height:2px;
      border-radius:999px;
      background:linear-gradient(90deg, rgba(122,36,56,.42), rgba(176,74,100,.34), rgba(216,202,190,.28));
      box-shadow:0 0 10px rgba(122,36,56,.08);
      pointer-events:none;
    }}
    .mini-card {{
      padding:12px 14px;
      min-height:84px;
      display:flex;
      flex-direction:column;
      justify-content:center;
      align-items:center;
      text-align:center;
      gap:2px;
    }}
    .metric-card {{
      padding:14px 16px;
      min-height:94px;
      display:flex;
      flex-direction:column;
      justify-content:space-between;
    }}
    .metric-label {{ display:block; color:var(--muted); font-size:13px; line-height:1.45; font-weight:560; }}
    .metric-card strong {{ display:block; margin-top:8px; font-size:22px; line-height:1; font-weight:880; letter-spacing:-0.02em; font-variant-numeric: tabular-nums; color:#fff; text-shadow:0 2px 14px rgba(122,36,56,.12), 0 1px 0 rgba(255,255,255,.08); }}
    .metric-footer {{ margin-top:7px; color:rgba(211,222,235,.66); font-size:12px; line-height:1.35; }}
    .mini-card .metric-label {{ font-size:13px; line-height:1.35; }}
    .mini-card strong {{ display:block; margin-top:6px; font-size:20px; line-height:1.15; font-variant-numeric: tabular-nums; color:#fff; font-weight:850; text-shadow:0 2px 14px rgba(122,36,56,.10); }}
    .mini-card .metric-footer {{ font-size:12px; line-height:1.35; }}
    .profit-details-card {{ margin:0 0 14px; }}
    .details-panel {{ padding:16px; }}
    .compact-header {{ margin-bottom:12px; }}
    .details-grid {{ display:grid; grid-template-columns:repeat(5, minmax(0,1fr)); gap:10px; }}
    .detail-chip {{
      position:relative;
      overflow:hidden;
      min-height:68px;
      padding:12px;
      border-radius:10px;
      background:rgba(24,38,58,.76);
      border:1px solid rgba(150,169,196,.18);
      display:flex;
      flex-direction:column;
      justify-content:center;
      gap:6px;
    }}
    .detail-chip span {{ color:var(--muted); font-size:12px; font-weight:560; }}
    .detail-chip strong {{ color:#fff; font-size:18px; line-height:1.1; font-weight:850; font-variant-numeric:tabular-nums; text-shadow:0 2px 12px rgba(122,36,56,.09); }}
    .content-grid {{ display:grid; grid-template-columns:1.3fr .92fr; gap:14px; margin-bottom:14px; }}
    .panel {{ padding:18px; }}
    .panel-header {{ display:flex; align-items:flex-start; justify-content:space-between; gap:14px; margin-bottom:16px; }}
    .eyebrow {{ margin:0; color:rgba(238,245,252,.88); font-size:13px; font-weight:750; letter-spacing:.02em; }}
    .subtle {{ margin:6px 0 0; color:rgba(184,198,214,.78); font-size:13px; }}
    .chart-shell {{ height:300px; }}
    .combo-chart {{ width:100%; height:100%; display:block; }}
    .area-chart {{ width:100%; height:100%; display:block; }}
    .area-chart svg {{ width:100%; height:100%; display:block; overflow:visible; }}
    .area-value {{ fill:#fff; font-size:11px; font-weight:850; font-family:Inter, "Avenir Next", "Segoe UI", "Noto Sans Arabic", system-ui, sans-serif; }}
    .area-date {{ fill:rgba(255,255,255,.56); font-size:10px; font-weight:700; font-family:Inter, "Avenir Next", "Segoe UI", "Noto Sans Arabic", system-ui, sans-serif; }}
    .chart-line {{ filter: drop-shadow(0 2px 6px rgba(177,42,91,.18)); }}
    .chart-point {{ filter: drop-shadow(0 2px 5px rgba(0,0,0,.18)); }}
    .carrier-layout {{
      display:grid;
      grid-template-columns:156px 1fr;
      gap:16px;
      align-items:center;
    }}
    .carrier-donut {{
      width:min(156px, 48vw);
      aspect-ratio:1;
      border-radius:50%;
      position:relative;
      margin:0 auto;
      padding:0;
      background:linear-gradient(180deg, rgba(33,47,70,.94), rgba(17,29,46,.90));
      border:1px solid rgba(135,154,180,.12);
      box-shadow:0 20px 34px rgba(0,0,0,.34), 0 0 24px rgba(122,36,56,.05), inset 0 1px 0 rgba(255,255,255,.08);
    }}
    .carrier-donut canvas {{
      position:absolute;
      inset:0;
      width:100%;
      height:100%;
    }}
    .carrier-center {{
      position:absolute;
      inset:0;
      display:flex;
      align-items:center;
      justify-content:center;
      z-index:1;
      text-align:center;
      padding:18px;
      color:#fff;
      pointer-events:none;
    }}
    .carrier-center span {{ display:block; font-size:11px; color:rgba(255,255,255,.72); }}
    .carrier-center strong {{ display:block; font-size:22px; line-height:1; font-weight:850; }}
    .carrier-legend {{ display:grid; gap:10px; }}
    .carrier-row {{
      display:grid;
      grid-template-columns:12px 1fr auto;
      gap:10px;
      align-items:center;
      padding:12px 14px;
      border-radius:10px;
      border:1px solid rgba(150,169,196,.18);
      background:rgba(24,38,58,.72);
    }}
    .carrier-dot {{ width:12px; height:12px; border-radius:50%; box-shadow:0 0 0 4px rgba(122,36,56,.045); }}
    .carrier-name {{ font-size:13px; font-weight:700; color:var(--text); }}
    .carrier-meta {{ font-size:12px; color:rgba(220,231,242,.70); white-space:nowrap; }}
    .table-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
    .table-wrap {{ overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; }}
    th, td {{
      padding:12px 10px;
      border-bottom:1px solid rgba(150,169,196,.14);
      text-align:right;
      white-space:nowrap;
    }}
    th {{
      color:rgba(184,198,214,.82);
      font-size:12px;
      font-weight:700;
    }}
    td {{
      color:rgba(244,248,252,.92);
      font-size:13px;
    }}
    .rank {{
      width:36px;
      color:rgba(202,215,229,.66);
      font-variant-numeric: tabular-nums;
    }}
    .table-card-title {{
      display:flex;
      align-items:flex-start;
      justify-content:space-between;
      gap:12px;
      margin-bottom:14px;
    }}
    .table-card-title p {{ margin:6px 0 0; color:var(--muted); font-size:12px; }}
    .statement-summary {{ display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:12px; margin-bottom:14px; }}
    .statement-summary .mini-card strong {{ font-size:18px; }}
    .statement-panel {{ margin-top:16px; }}
    .finance-card strong {{ font-size:18px; }}
    .finance-total-card strong {{ font-size:16px; }}
    .breakdown-line {{ color:rgba(255,255,255,.70); font-size:12px; line-height:1.5; margin-top:2px; }}
    .footer-note {{ margin-top:16px; color:rgba(255,255,255,.48); font-size:12px; }}

    .login-screen {{
      min-height:100vh;
      background:radial-gradient(circle at top, rgba(122,36,56,.22), transparent 35%), #0F1113;
      display:flex;
      align-items:center;
      justify-content:center;
      padding:24px;
      color:#F5F2EE;
    }}
    .login-screen.is-hidden {{ display:none; }}
    .login-screen.is-hiding {{ opacity:0; transition:opacity .25s ease; }}
    .login-card {{
      width:100%;
      max-width:420px;
      background:rgba(23,26,31,.92);
      border:1px solid #252932;
      border-radius:20px;
      box-shadow:0 24px 80px rgba(0,0,0,.35);
      padding:32px;
      color:#F5F2EE;
      opacity:0;
      transform:translateY(12px) scale(.98);
      animation:loginCardIn .35s ease-out forwards;
    }}
    @keyframes loginCardIn {{ to {{ opacity:1; transform:translateY(0) scale(1); }} }}
    .login-brand {{ display:flex; flex-direction:column; align-items:center; gap:10px; margin-bottom:18px; text-align:center; }}
    .login-brand img {{ width:132px; height:auto; filter:none; }}
    .login-card h2 {{ margin:8px 0 4px; font-size:24px; line-height:1.2; }}
    .login-card p {{ margin:0; color:#A7A7A7; font-size:13px; line-height:1.7; }}
    .login-form {{ display:grid; gap:14px; margin-top:22px; }}
    .login-field {{ display:grid; gap:8px; }}
    .login-field label {{ color:#D8CABE; font-size:13px; font-weight:650; }}
    .login-input-wrap {{ position:relative; }}
    .login-card input {{
      width:100%;
      background:#0F1113;
      border:1px solid #252932;
      border-radius:14px;
      color:#F5F2EE;
      padding:14px 16px;
      outline:none;
      font:inherit;
    }}
    .login-card input:focus {{ border-color:#7A2438; box-shadow:0 0 0 4px rgba(122,36,56,.18); }}
    .password-toggle {{
      position:absolute;
      left:8px;
      top:50%;
      transform:translateY(-50%);
      width:auto !important;
      min-height:34px;
      padding:0 10px !important;
      border-radius:10px !important;
      background:transparent !important;
      border:1px solid #252932 !important;
      color:#A7A7A7 !important;
      font-size:12px;
    }}
    .login-button {{
      width:100%;
      background:#7A2438;
      color:#F5F2EE;
      border:none;
      border-radius:14px;
      padding:14px 16px;
      cursor:pointer;
      font-weight:700;
      font:inherit;
    }}
    .login-button:hover {{ background:#B04A64; }}
    .login-error {{
      display:none;
      color:#C86B6B;
      background:rgba(200,107,107,.10);
      border:1px solid rgba(200,107,107,.20);
      border-radius:12px;
      padding:10px 12px;
      margin-top:12px;
      font-size:14px;
    }}
    .login-error.is-visible {{ display:block; }}
    .dashboard-shell {{ opacity:0; transform:scale(.98); pointer-events:none; }}
    .dashboard-shell.is-visible {{ opacity:1; transform:scale(1); pointer-events:auto; transition:opacity .35s ease-out, transform .35s ease-out; }}
    .hero, .metric-card, .panel, .mini-card, .detail-chip, .carrier-row, .carrier-donut {{
      background:var(--card);
      border-color:var(--border);
      box-shadow:var(--shadow);
      backdrop-filter:none;
      -webkit-backdrop-filter:none;
    }}
    .hero {{ border-radius:18px; min-height:78px; }}
    .brand-identity {{ display:grid; gap:2px; justify-items:start; }}
    .brand-identity strong {{ color:var(--text); font-size:14px; line-height:1.1; letter-spacing:.02em; }}
    .brand-identity span {{ color:var(--muted); font-size:11px; line-height:1.2; }}
    .subhero-tools {{ display:flex; align-items:center; justify-content:flex-end; gap:10px; flex-wrap:wrap; }}
    .theme-toggle, .logout-button {{
      min-height:38px;
      border-radius:12px;
      border:1px solid var(--border);
      background:var(--sidebar);
      color:var(--text);
      padding:0 12px;
      font:inherit;
      font-size:12px;
      font-weight:700;
      cursor:pointer;
    }}
    .logout-button {{ color:var(--muted); }}
    .primary-button {{
      background:var(--accent);
      border-color:var(--accent);
      color:#F5F2EE;
      box-shadow:0 10px 22px rgba(122,36,56,.14);
    }}
    .primary-button:hover {{ background:var(--accent-light); filter:none; }}
    .badge {{ background:var(--sidebar); border-color:var(--border); color:var(--muted); }}
    .metric-card::before, .panel::before, .mini-card::before {{ background:none; }}
    .metric-card::after, .mini-card::after, .detail-chip::after {{
      background:linear-gradient(90deg, var(--accent), var(--accent-light), var(--beige));
      opacity:.55;
      box-shadow:none;
    }}
    .hero-title h1, .metric-card strong, .mini-card strong, .detail-chip strong, .carrier-center strong {{ color:var(--text); text-shadow:none; }}
    .hero-title p, .metric-footer, .subtle, .carrier-meta, .breakdown-line, th, .rank {{ color:var(--muted); }}
    .eyebrow, td, .carrier-name {{ color:var(--text); }}
    th, td {{ border-bottom-color:var(--grid); }}
    .carrier-center {{ color:var(--text); }}
    .carrier-center span {{ color:var(--muted); }}
    .dashboard-footer {{
      margin-top:18px;
      padding:16px 0 2px;
      border-top:1px solid var(--border);
      color:var(--muted);
      text-align:center;
      font-size:12px;
      line-height:1.7;
    }}
    .dashboard-footer strong {{ display:block; color:var(--muted); font-size:13px; font-weight:700; }}
    .missing-seq {{
      color:var(--text);
      background:var(--sidebar);
      border:1px solid var(--border);
      border-radius:14px;
      padding:14px 16px;
      line-height:1.9;
      font-size:14px;
      word-break:break-word;
    }}
    body:not(.dark) .login-screen {{ background:#0F1113; }}

    @media (max-width: 1100px) {{
      .metric-grid, .content-grid, .table-grid, .mini-band, .finance-band, .statement-summary, .details-grid {{ grid-template-columns:1fr; }}
      .carrier-layout {{ grid-template-columns:1fr; }}
      .hero {{ padding:12px 14px; margin-bottom:8px; align-items:center; gap:12px; }}
      .brand-lockup {{ flex:0 1 auto; min-width:0; }}
      .brand-logo {{ width:min(82px, 24vw); }}
      .brand-divider {{ height:32px; margin-inline:7px; }}
      .hero-title {{ position:relative; inset:auto; transform:none; min-width:0; padding:0; align-items:flex-end; width:auto; flex:0 1 auto; }}
      .hero-copy span {{ font-size:12px; }}
      .hero-title h1 {{ font-size:16px; }}
      .hero-title p {{ font-size:10px; }}
      .metric-card {{ min-height:88px; }}
      .mini-card {{ min-height:78px; }}
      .detail-chip {{ min-height:64px; }}
      .subhero-tools {{ width:100%; justify-content:flex-start; }}
      .theme-toggle, .logout-button {{ flex:1 1 auto; }}
    }}
  </style>
</head>
<body class="dark">
  <script>
    (function () {{
      const savedTheme = localStorage.getItem('gfTheme') || 'dark';
      document.body.classList.toggle('dark', savedTheme !== 'day');
    }})();
  </script>
  <section class="login-screen" id="loginScreen" aria-label="تسجيل الدخول">
    <div class="login-card">
      <div class="login-brand">
        <img src="gf_logo_mark_clean.png" alt="GF" />
        <div>
          <h2>تسجيل الدخول</h2>
          <p>أدخل بيانات الدخول للوصول إلى لوحة التحكم</p>
        </div>
      </div>
      <form class="login-form" id="loginForm">
        <div class="login-field">
          <label for="loginUsername">اسم المستخدم</label>
          <input id="loginUsername" name="username" autocomplete="username" required />
        </div>
        <div class="login-field">
          <label for="loginPassword">كلمة المرور</label>
          <div class="login-input-wrap">
            <input id="loginPassword" name="password" type="password" autocomplete="current-password" required />
            <button class="password-toggle" id="passwordToggle" type="button">إظهار</button>
          </div>
        </div>
        <button class="login-button" type="submit">دخول</button>
        <div class="login-error" id="loginError">بيانات الدخول غير صحيحة</div>
      </form>
    </div>
  </section>
  <main class="page dashboard-shell" id="dashboardShell" aria-hidden="true">
    <header class="hero">
      <div class="hero-title">
        <h1>لوحة التحكم المالية المختصرة</h1>
        <p>نظرة عامة على أداء الأعمال</p>
      </div>
      <div class="brand-lockup">
        <img class="brand-logo" src="gf_logo_mark_clean.png" alt="GF Smart Accounting Solutions" />
        <div class="brand-divider" aria-hidden="true"></div>
        <div class="hero-copy brand-identity">
          <strong>GF Analytics</strong>
          <span>Financial Intelligence Dashboard</span>
        </div>
      </div>
    </header>

    <div class="subhero">
      <div class="hero-tools">
        <span class="badge">{period_label}</span>
        <a class="primary-button" href="{details_href}"><span>فتح ملف Excel</span></a>
      </div>
      <div class="subhero-tools">
        <button class="theme-toggle" id="themeToggle" type="button">🌙 Night Mode</button>
        <button class="logout-button" id="logoutButton" type="button">تسجيل خروج</button>
      </div>
    </div>

    <section class="metric-grid">
      {metric_card("إجمالي الإيرادات", fmt_money(totals["revenue"]), "إجمالي ما دفعه العملاء")}
      {metric_card("إجمالي الربح", fmt_money(totals["total"]), f"الشحنات المحسوبة: {period_label}")}
      {metric_card("عدد الشحنات", totals["active"], f"غير داخلة في الربح: {totals['excluded']}")}
      {metric_card("مبلغ COD", fmt_money(totals["cod_amount"]), "إجمالي مبالغ COD")}
    </section>

    {profit_details}

    <section class="content-grid">
      <article class="panel">
        <div class="panel-header">
          <div>
            <p class="eyebrow">الأداء اليومي</p>
            <p class="subtle">آخر 7 أيام</p>
          </div>
        </div>
        <div class="chart-shell">
          <canvas id="dailyComboChart" class="combo-chart" aria-label="الأداء اليومي"></canvas>
        </div>
      </article>

      <article class="panel">
        <div class="panel-header">
          <div>
            <p class="eyebrow">شركات الشحن</p>
            <p class="subtle">نسبة الربح لكل ناقل</p>
          </div>
        </div>
        <div class="carrier-layout">
          <div class="carrier-donut">
            <canvas id="carrierDonutChart" aria-label="شركات الشحن"></canvas>
            <div class="carrier-center">
              <div>
                <strong>{len(top_carriers)}</strong>
                <span>شركات</span>
              </div>
            </div>
          </div>
          <div class="carrier-legend">
            {carrier_legend_html}
          </div>
        </div>
      </article>
    </section>

    <section class="table-grid">
      <article class="panel">
        <div class="table-card-title">
          <div>
            <p class="eyebrow">أفضل العملاء</p>
            <p class="subtle">حسب إجمالي الربح</p>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th class="rank">#</th><th>العميل</th><th>عدد</th><th>الربح</th></tr>
            </thead>
            <tbody>{merchant_rows}</tbody>
          </table>
        </div>
      </article>

      <article class="panel">
        <div class="table-card-title">
          <div>
            <p class="eyebrow">أفضل المدن</p>
            <p class="subtle">حسب إجمالي الربح</p>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th class="rank">#</th><th>المدينة</th><th>عدد</th><th>الربح</th></tr>
            </thead>
            <tbody>{city_rows}</tbody>
          </table>
        </div>
      </article>
    </section>

    <section class="panel" style="margin-top:16px;">
      <div class="panel-header">
        <div>
          <p class="eyebrow">العمليات المالية</p>
          <p class="subtle">إيداعات ومخصومات واستردادات</p>
        </div>
      </div>
      {finance_cards}
    </section>

    <section class="panel statement-panel">
      <div class="panel-header">
        <div>
          <p class="eyebrow">كشف الحساب الجاري</p>
          <p class="subtle">ملخص الحركات من الشيت</p>
        </div>
      </div>
      <div class="statement-summary">
        <div class="mini-card quick-card">
          <span class="metric-label">إجمالي الإيداعات</span>
          <strong>{fmt_money(statement["summary"].get("deposits_total", 0))}</strong>
        </div>
        <div class="mini-card quick-card">
          <span class="metric-label">إجمالي المصاريف</span>
          <strong>{fmt_money(statement["summary"].get("expenses_total", 0))}</strong>
        </div>
        <div class="mini-card quick-card">
          <span class="metric-label">رسوم الحوالات</span>
          <strong>{fmt_money(statement["summary"].get("transfer_fees_total", 0))}</strong>
        </div>
        <div class="mini-card quick-card">
          <span class="metric-label">صافي الحركة</span>
          <strong>{fmt_money(statement["summary"].get("net_total", 0))}</strong>
        </div>
        <div class="mini-card quick-card">
          <span class="metric-label">عدد المصاريف</span>
          <strong>{int(statement["summary"].get("expenses_count", 0))}</strong>
        </div>
      </div>
      <div class="table-wrap statement-table">
        <table>
          <thead>
            <tr><th>التاريخ</th><th>نوع المصروف</th><th>المبلغ</th></tr>
          </thead>
          <tbody>{statement_rows}</tbody>
        </table>
      </div>
    </section>

    <section class="panel" style="margin-top:16px;">
      <div class="panel-header">
        <div>
          <p class="eyebrow">طلبات COD</p>
          <p class="subtle">كل طلب على حدة</p>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>رقم الطلب</th>
              <th>التاجر</th>
              <th>مبلغ COD</th>
              <th>تاريخ الطلب</th>
              <th>تاريخ التحصل</th>
              <th>تاريخ التحويل</th>
            </tr>
          </thead>
          <tbody>{cod_rows}</tbody>
        </table>
      </div>
    </section>

    <section class="panel" style="margin-top:16px;">
      <div class="panel-header">
        <div>
          <p class="eyebrow">الأرقام المتسلسلة الناقصة</p>
          <p class="subtle">الفجوات بين أول وآخر رقم في الشحنات</p>
        </div>
      </div>
      <div class="missing-seq">{missing_sequence_text}</div>
    </section>

    <footer class="dashboard-footer">
      <strong>GF Smart Accounting Solutions</strong>
      <span>Last Updated: {last_updated}</span>
    </footer>
  </main>
  <script>

    const loginScreen = document.getElementById('loginScreen');
    const dashboardShell = document.getElementById('dashboardShell');
    const loginForm = document.getElementById('loginForm');
    const loginError = document.getElementById('loginError');
    const passwordInput = document.getElementById('loginPassword');
    const passwordToggle = document.getElementById('passwordToggle');
    const themeToggle = document.getElementById('themeToggle');
    const logoutButton = document.getElementById('logoutButton');

    function setDashboardVisible(visible) {{
      if (visible) {{
        loginScreen.classList.add('is-hidden');
        dashboardShell.classList.add('is-visible');
        dashboardShell.setAttribute('aria-hidden', 'false');
      }} else {{
        loginScreen.classList.remove('is-hidden', 'is-hiding');
        dashboardShell.classList.remove('is-visible');
        dashboardShell.setAttribute('aria-hidden', 'true');
      }}
    }}

    function completeLogin() {{
      localStorage.setItem('gfLeadAuthenticated', 'true');
      loginScreen.classList.add('is-hiding');
      setTimeout(() => setDashboardVisible(true), 250);
    }}

    function applyTheme(theme) {{
      const isDark = theme !== 'day';
      document.body.classList.toggle('dark', isDark);
      localStorage.setItem('gfTheme', isDark ? 'dark' : 'day');
      themeToggle.textContent = isDark ? '☀️ Day Mode' : '🌙 Night Mode';
    }}

    applyTheme(localStorage.getItem('gfTheme') || 'dark');
    if (localStorage.getItem('gfLeadAuthenticated') === 'true') setDashboardVisible(true);

    passwordToggle.addEventListener('click', () => {{
      const showing = passwordInput.type === 'text';
      passwordInput.type = showing ? 'password' : 'text';
      passwordToggle.textContent = showing ? 'إظهار' : 'إخفاء';
    }});

    loginForm.addEventListener('submit', async (event) => {{
      event.preventDefault();
      loginError.classList.remove('is-visible');
      const payload = {{
        username: document.getElementById('loginUsername').value,
        password: passwordInput.value
      }};
      try {{
        const response = await fetch('/api/login', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify(payload)
        }});
        if (!response.ok) throw new Error('invalid');
        completeLogin();
      }} catch (error) {{
        loginError.classList.add('is-visible');
      }}
    }});

    themeToggle.addEventListener('click', () => {{
      applyTheme(document.body.classList.contains('dark') ? 'day' : 'dark');
    }});

    logoutButton.addEventListener('click', async () => {{
      localStorage.removeItem('gfLeadAuthenticated');
      try {{ await fetch('/api/logout', {{ method: 'POST' }}); }} catch (error) {{}}
      setDashboardVisible(false);
    }});

    const dailyLabels = {json.dumps(daily_chart_labels, ensure_ascii=False)};
    const dailyProfit = {json.dumps(daily_chart_profit, ensure_ascii=False)};
    const dailyRevenue = {json.dumps(daily_chart_revenue, ensure_ascii=False)};
    const dailyCount = {json.dumps(daily_chart_count, ensure_ascii=False)};
    const carrierLabels = {json.dumps(carrier_chart_labels, ensure_ascii=False)};
    const carrierValues = {json.dumps(carrier_chart_values, ensure_ascii=False)};
    const carrierColors = {json.dumps(carrier_chart_colors, ensure_ascii=False)};
    const chartCanvas = document.getElementById('dailyComboChart');
    const carrierCanvas = document.getElementById('carrierDonutChart');
    function drawCanvasComboChart(canvas, labels, profit, revenue, counts) {{
      const ctx = canvas.getContext('2d');
      const dpr = window.devicePixelRatio || 1;
      const rect = canvas.getBoundingClientRect();
      canvas.width = Math.max(1, Math.floor(rect.width * dpr));
      canvas.height = Math.max(1, Math.floor(rect.height * dpr));
      ctx.setTransform(dpr, 0, 0, dpr, 0, 0);
      const width = rect.width;
      const height = rect.height;
      const pad = {{ top: 34, right: 36, bottom: 34, left: 44 }};
      const plotW = width - pad.left - pad.right;
      const plotH = height - pad.top - pad.bottom;
      const maxBars = Math.max(...profit, ...revenue, 1);
      const maxCount = Math.max(...counts, 1);

      ctx.clearRect(0, 0, width, height);
      ctx.font = '10px Inter, "Noto Sans Arabic", system-ui';
      ctx.textAlign = 'center';
      ctx.textBaseline = 'middle';

      for (let i = 0; i <= 4; i += 1) {{
        const y = pad.top + (plotH * i / 4);
        ctx.strokeStyle = 'rgba(255,255,255,.06)';
        ctx.lineWidth = 1;
        ctx.beginPath();
        ctx.moveTo(pad.left, y);
        ctx.lineTo(width - pad.right, y);
        ctx.stroke();
      }}

      const step = plotW / Math.max(labels.length, 1);
      const barW = Math.min(18, step * 0.27);
      const profitGradient = ctx.createLinearGradient(0, pad.top, 0, height - pad.bottom);
      profitGradient.addColorStop(0, 'rgba(216,202,190,.88)');
      profitGradient.addColorStop(1, 'rgba(216,202,190,.20)');
      const revenueGradient = ctx.createLinearGradient(0, pad.top, 0, height - pad.bottom);
      revenueGradient.addColorStop(0, 'rgba(176,74,100,.86)');
      revenueGradient.addColorStop(1, 'rgba(176,74,100,.20)');

      labels.forEach((label, index) => {{
        const centerX = pad.left + step * index + step / 2;
        const profitH = (profit[index] / maxBars) * plotH;
        const revenueH = (revenue[index] / maxBars) * plotH;
        const profitX = centerX - barW - 2;
        const revenueX = centerX + 2;
        ctx.fillStyle = profitGradient;
        ctx.beginPath();
        ctx.roundRect(profitX, height - pad.bottom - profitH, barW, profitH, 5);
        ctx.fill();
        ctx.fillStyle = revenueGradient;
        ctx.beginPath();
        ctx.roundRect(revenueX, height - pad.bottom - revenueH, barW, revenueH, 5);
        ctx.fill();
        ctx.fillStyle = 'rgba(246,251,255,.58)';
        ctx.fillText(label, centerX, height - 16);
      }});

      const linePoints = counts.map((count, index) => {{
        const centerX = pad.left + step * index + step / 2;
        const y = height - pad.bottom - ((count / maxCount) * plotH);
        return [centerX, y];
      }});
      ctx.strokeStyle = 'rgba(122,36,56,.95)';
      ctx.lineWidth = 2.5;
      ctx.beginPath();
      linePoints.forEach(([x, y], index) => {{
        if (index === 0) ctx.moveTo(x, y);
        else ctx.lineTo(x, y);
      }});
      ctx.stroke();
      linePoints.forEach(([x, y]) => {{
        ctx.fillStyle = 'rgba(122,36,56,.95)';
        ctx.beginPath();
        ctx.arc(x, y, 4, 0, Math.PI * 2);
        ctx.fill();
        ctx.strokeStyle = '#0F1113';
        ctx.lineWidth = 2;
        ctx.stroke();
      }});

      ctx.textAlign = 'right';
      ctx.fillStyle = 'rgba(246,251,255,.68)';
      ctx.fillText('صافي الربح', width - pad.right, 16);
      ctx.fillStyle = 'rgba(176,74,100,.78)';
      ctx.fillText('الإيراد', width - pad.right - 82, 16);
      ctx.fillStyle = 'rgba(122,36,56,.78)';
      ctx.fillText('عدد الطلبات', width - pad.right - 142, 16);
    }}

    function drawCanvasDonutChart(canvas, values, colors) {{
      const ctx = canvas.getContext('2d');
      const dpr = window.devicePixelRatio || 1;
      const rect = canvas.getBoundingClientRect();
      canvas.width = Math.max(1, Math.floor(rect.width * dpr));
      canvas.height = Math.max(1, Math.floor(rect.height * dpr));
      ctx.setTransform(dpr, 0, 0, dpr, 0, 0);
      const cx = rect.width / 2;
      const cy = rect.height / 2;
      const radius = Math.min(rect.width, rect.height) / 2 - 8;
      const inner = radius * .56;
      const total = values.reduce((sum, value) => sum + value, 0) || 1;
      let start = -Math.PI / 2;

      ctx.clearRect(0, 0, rect.width, rect.height);
      values.forEach((value, index) => {{
        const end = start + (value / total) * Math.PI * 2;
        ctx.beginPath();
        ctx.arc(cx, cy, radius, start, end);
        ctx.arc(cx, cy, inner, end, start, true);
        ctx.closePath();
        ctx.fillStyle = colors[index % colors.length];
        ctx.fill();
        ctx.strokeStyle = 'rgba(7,17,29,.78)';
        ctx.lineWidth = 2;
        ctx.stroke();
        start = end;
      }});

      ctx.beginPath();
      ctx.arc(cx, cy, inner - 1, 0, Math.PI * 2);
      ctx.fillStyle = 'rgba(5,8,14,.94)';
      ctx.fill();
      ctx.strokeStyle = 'rgba(255,255,255,.08)';
      ctx.lineWidth = 1;
      ctx.stroke();
    }}

    if (chartCanvas && window.Chart) {{
      const chartContext = chartCanvas.getContext('2d');
      const profitGradient = chartContext.createLinearGradient(0, 0, 0, 280);
      profitGradient.addColorStop(0, 'rgba(216,202,190,0.88)');
      profitGradient.addColorStop(1, 'rgba(216,202,190,0.20)');
      const revenueGradient = chartContext.createLinearGradient(0, 0, 0, 280);
      revenueGradient.addColorStop(0, 'rgba(176,74,100,0.86)');
      revenueGradient.addColorStop(1, 'rgba(176,74,100,0.20)');
      new Chart(chartCanvas, {{
        data: {{
          labels: dailyLabels,
          datasets: [
            {{
              type: 'bar',
              label: 'صافي الربح',
              data: dailyProfit,
              yAxisID: 'money',
              backgroundColor: profitGradient,
              borderColor: 'rgba(216,202,190,.95)',
              borderWidth: 1,
              borderRadius: 6,
              maxBarThickness: 20
            }},
            {{
              type: 'bar',
              label: 'الإيراد',
              data: dailyRevenue,
              yAxisID: 'money',
              backgroundColor: revenueGradient,
              borderColor: 'rgba(176,74,100,.96)',
              borderWidth: 1,
              borderRadius: 6,
              maxBarThickness: 20
            }},
            {{
              type: 'line',
              label: 'عدد الطلبات',
              data: dailyCount,
              yAxisID: 'count',
              borderColor: 'rgba(122,36,56,.95)',
              backgroundColor: 'rgba(122,36,56,0.12)',
              borderWidth: 2.5,
              tension: 0.38,
              pointRadius: 3.5,
              pointHoverRadius: 5,
              pointBackgroundColor: 'rgba(122,36,56,.95)',
              pointBorderColor: '#0F1113',
              pointBorderWidth: 2
            }}
          ]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          interaction: {{ mode: 'index', intersect: false }},
          plugins: {{
            legend: {{
              position: 'top',
              align: 'end',
              labels: {{
                color: 'rgba(246,251,255,.76)',
                boxWidth: 10,
                boxHeight: 10,
                usePointStyle: true,
                font: {{ family: 'Inter, "Noto Sans Arabic", system-ui', size: 11 }}
              }}
            }},
            tooltip: {{
              rtl: true,
              textDirection: 'rtl',
              backgroundColor: 'rgba(5,8,14,.94)',
              borderColor: 'rgba(255,255,255,.14)',
              borderWidth: 1,
              titleColor: '#fff',
              bodyColor: 'rgba(246,251,255,.82)'
            }}
          }},
          scales: {{
            x: {{
              grid: {{ color: 'rgba(255,255,255,.05)' }},
              ticks: {{ color: 'rgba(246,251,255,.58)', font: {{ size: 10 }} }}
            }},
            money: {{
              position: 'left',
              grid: {{ color: 'rgba(255,255,255,.06)' }},
              ticks: {{ color: 'rgba(246,251,255,.58)', font: {{ size: 10 }} }}
            }},
            count: {{
              position: 'right',
              grid: {{ drawOnChartArea: false }},
              ticks: {{ color: 'rgba(122,36,56,.72)', precision: 0, font: {{ size: 10 }} }}
            }}
          }}
        }}
      }});
    }} else if (chartCanvas) {{
      drawCanvasComboChart(chartCanvas, dailyLabels, dailyProfit, dailyRevenue, dailyCount);
    }}

    if (carrierCanvas && window.Chart) {{
      new Chart(carrierCanvas, {{
        type: 'doughnut',
        data: {{
          labels: carrierLabels,
          datasets: [{{
            data: carrierValues,
            backgroundColor: carrierColors,
            borderColor: 'rgba(7,17,29,.78)',
            borderWidth: 2,
            hoverOffset: 3
          }}]
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          cutout: '58%',
          plugins: {{
            legend: {{ display: false }},
            tooltip: {{
              rtl: true,
              textDirection: 'rtl',
              backgroundColor: 'rgba(5,8,14,.94)',
              borderColor: 'rgba(255,255,255,.14)',
              borderWidth: 1,
              titleColor: '#fff',
              bodyColor: 'rgba(246,251,255,.82)'
            }}
          }}
        }}
      }});
    }} else if (carrierCanvas) {{
      drawCanvasDonutChart(carrierCanvas, carrierValues, carrierColors);
    }}
  </script>
</body>
</html>"""


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    if db_enabled():
        try:
            totals, top_merchants, top_cities, top_carriers, statuses, daily, daily_revenue, daily_count, finance, cod_items, return_items, statement, period_label, missing_sequence_numbers = load_data_from_db()
        except Exception:
            totals, top_merchants, top_cities, top_carriers, statuses, daily, daily_revenue, daily_count, finance, cod_items, return_items, statement, period_label, missing_sequence_numbers = load_data()
    else:
        totals, top_merchants, top_cities, top_carriers, statuses, daily, daily_revenue, daily_count, finance, cod_items, return_items, statement, period_label, missing_sequence_numbers = load_data()
    data = {
        "totals": totals,
        "top_merchants": top_merchants,
        "top_cities": top_cities,
        "daily": daily,
        "daily_revenue": daily_revenue,
        "daily_count": daily_count,
        "finance": finance,
        "cod_items": cod_items,
        "return_items": return_items,
        "top_carriers": top_carriers,
        "statement": statement,
        "period_label": period_label,
        "missing_sequence_numbers": missing_sequence_numbers,
    }
    with open(INDEX, "w", encoding="utf-8") as f:
        f.write(build_html(data))
    print(INDEX)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        traceback.print_exc()
        raise
